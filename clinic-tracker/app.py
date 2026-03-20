import os
import csv
import sqlite3
import json
import re
from datetime import datetime, timedelta
from functools import wraps
from io import StringIO, BytesIO
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, g, send_from_directory, Response)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'vip-medical-dev-secret-2024')
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB
DATABASE = os.path.join(os.path.dirname(__file__), 'clinic_tracker.db')

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx', 'txt'}
ALLOWED_DOC_EXTENSIONS = {'pdf', 'xlsx', 'docx', 'doc', 'xls'}

DEPARTMENTS = [
    'Strategic Growth', 'Marketing', 'IT', 'Inventory', 'Operations',
    'Accounting / Accounts Payable', 'Credentialing', 'RODs & Clinic Leads',
    'Malpractice', 'Special Operations', 'Referrals Outreach', 'HR'
]

TIME_PHASES = [
    'Scouting', 'After Lease Signed', '2 Months Before Opening', '1 Month Before Opening',
    '2 Weeks Before Opening', '1 Week Before Opening', 'Week Before Opening', 'Opening Day',
    '1 Week After Opening', '1 Month After Opening', '2 Months After Opening',
    '3 Months After Opening', 'When New Provider Hired'
]
STATUSES = ['Not Started', 'In Progress', 'Complete', 'Blocked']

SUPPLY_ORDER_STATUSES = ['Not Ordered', 'Ordered', 'Delivered', 'Issue']

CONSTRUCTION_STATUSES = ['Not Started', 'In Progress', 'Complete', 'Delayed']

# Master supply items list: (tab, category, item_name, default_qty, team_ordering)
SUPPLY_MASTER = [
    # OFFICE - HIGH PRIORITY
    ('Office', 'HIGH PRIORITY', 'Front Desk Desktop', '1', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Trash Bags', '2 boxes', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Contractor-sized trash bags', '2 rolls', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Tool Box', '1', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Small Electric Screwdriver', '1', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Regular iPads', '3', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Scanner', '1', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Printer', '1', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Scissors', '2', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Packing tape', '1 roll', 'Inventory'),
    ('Office', 'HIGH PRIORITY', 'Wire Storage Shelves', '3', 'Inventory'),
    # OFFICE - Office Supplies
    ('Office', 'Office Supplies', 'Nextech Credit Card Terminal', '1', 'Strategic Growth'),
    ('Office', 'Office Supplies', 'Wireless Headset', '4', 'Inventory'),
    ('Office', 'Office Supplies', "5' Ethernet Cords", '5-pack', 'Inventory'),
    ('Office', 'Office Supplies', 'Extension Cords', '4', 'Inventory'),
    ('Office', 'Office Supplies', 'iPad Covers', '3', 'Inventory'),
    ('Office', 'Office Supplies', 'iPad Charging Station', '1', 'Inventory'),
    ('Office', 'Office Supplies', 'Short Charging Cords for Station', '1 pack', 'Inventory'),
    ('Office', 'Office Supplies', 'Paper Shredder', '1', 'Inventory'),
    ('Office', 'Office Supplies', 'Label Maker', '1', 'Inventory'),
    ('Office', 'Office Supplies', 'Safe', '1', 'Inventory'),
    ('Office', 'Office Supplies', 'Drawer Organizers multipack', '# of rooms', 'Inventory'),
    ('Office', 'Office Supplies', 'Drawer Organizers long', '# of rooms', 'Inventory'),
    ('Office', 'Office Supplies', 'Drawer Organizers wide', '# of rooms', 'Inventory'),
    ('Office', 'Office Supplies', 'Under-Cabinet Storage Bins', '# of rooms', 'Inventory'),
    ('Office', 'Office Supplies', 'Desk Organizer for Front Desk', '3', 'Inventory'),
    # MEDICAL
    ('Medical', 'Medical Equipment', 'Ultrasound Machine - email GE rep', '2', 'Strategic Growth'),
    ('Medical', 'Medical Equipment', 'Ultrasound Cart', '2', 'Strategic Growth'),
    ('Medical', 'Medical Equipment', 'GE Software Scan Assist and Presets', '-', 'Strategic Growth'),
    ('Medical', 'Medical Equipment', 'Hill Medical Chairs', '# of rooms', 'Strategic Growth'),
    ('Medical', 'Medical Equipment', 'Medtronic Generator', '2', 'Inventory'),
    ('Medical', 'Medical Equipment', 'Medtronic Pump', '2', 'Inventory'),
    ('Medical', 'Medical Equipment', 'Stools with Back', '3', 'Inventory'),
    ('Medical', 'Medical Equipment', 'Step Stools', '2', 'Inventory'),
    ('Medical', 'Medical Equipment', 'Mayo Stands', '3', 'Inventory'),
    ('Medical', 'Medical Equipment', 'Ultrasound Gel Warmers', '# of rooms', 'Inventory'),
    ('Medical', 'Medical Equipment', 'Sharps Containers', '# of rooms', 'Inventory'),
    ('Medical', 'Medical Equipment', 'Glove Box Holders', '# of rooms', 'Inventory'),
    ('Medical', 'Medical Equipment', 'Pillows', '# of rooms', 'Inventory'),
    ('Medical', 'Medical Supplies', 'Pillow Cases', '2 Boxes', 'Inventory'),
    ('Medical', 'Medical Supplies', 'Cavi-Wipes', '12 Canisters', 'Inventory'),
    ('Medical', 'Medical Supplies', 'Amazon Shorts 200 Unisize', '200', 'Inventory'),
    ('Medical', 'Medical Supplies', 'McKesson Shorts 100 XL 100 XXXL', '200', 'Inventory'),
    ('Medical', 'Medical Supplies', 'Gowns', '1 case', 'Inventory'),
    ('Medical', 'Medical Supplies', 'Ultrasound Gel', '1 case', 'Inventory'),
    ('Medical', 'Medical Supplies', 'Ultrasound Gel Bottles', '# of rooms', 'Inventory'),
    ('Medical', 'Medical Supplies', 'Table paper', '2 boxes', 'Inventory'),
    ('Medical', 'Medical Supplies', '3cc syringes', '2 Boxes', 'Inventory'),
    ('Medical', 'Medical Supplies', '5cc syringes', '2 Boxes', 'Inventory'),
    ('Medical', 'Medical Supplies', '10cc syringes', '2 Boxes', 'Inventory'),
    ('Medical', 'Medical Supplies', '30cc syringes', '2 Boxes', 'Inventory'),
    ('Medical', 'Medical Supplies', 'Slip Tip Syringes', '2 Boxes', 'Inventory'),
    ('Medical', 'Medical Supplies', '18G Blunt Fill Needles', '6 Boxes', 'Inventory'),
    ('Medical', 'Medical Supplies', '25G 1.5 Needles', '6 Boxes', 'Inventory'),
    # IT
    ('IT', 'IT Equipment', 'Dream Machine Pro', '', 'IT'),
    ('IT', 'IT Equipment', 'PRO 24 POE', '', 'IT'),
    ('IT', 'IT Equipment', 'U6 Pro', '', 'IT'),
    ('IT', 'IT Equipment', 'G5 Dome Ultra', '', 'IT'),
    ('IT', 'IT Equipment', 'UPS Battery Back-Up', '', 'IT'),
    ('IT', 'IT Equipment', 'Disco Western Digital 4TB', '', 'IT'),
    ('IT', 'IT Equipment', 'Sonos POE Speakers', '', 'IT'),
    ('IT', 'IT Equipment', 'POE Speaker Mounts', '', 'IT'),
    ('IT', 'IT Equipment', 'Patient Room TVs', '', 'IT'),
    ('IT', 'IT Equipment', 'Staff Area TV', '', 'IT'),
    ('IT', 'IT Equipment', 'Waiting Room TV', '', 'IT'),
    ('IT', 'IT Equipment', 'TV Mounts', '', 'IT'),
    ('IT', 'IT Equipment', 'TV Corner Mounts', '', 'IT'),
    ('IT', 'IT Equipment', 'Access G3 Reader Pro', '', 'IT'),
    ('IT', 'IT Equipment', 'Access Cards', '', 'IT'),
    ('IT', 'IT Equipment', 'Access Door Hub Mini', '', 'IT'),
    ('IT', 'IT Equipment', '10G Direct Attach Cable', '', 'IT'),
    ('IT', 'IT Equipment', 'SFP to RJ45 Adapter', '', 'IT'),
    ('IT', 'IT Equipment', 'WiFi Smart Chime', '', 'IT'),
    ('IT', 'IT Equipment', 'Ethernet Cable Pack 10-Pack 10ft', '', 'IT'),
    # MARKETING - Bulk Order
    ('Marketing', 'Bulk Order', 'Business cards', '500', 'Marketing'),
    ('Marketing', 'Bulk Order', 'After Hours Cards', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Insurance cards', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Brochures Vein', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Brochures Pain', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Leg pain brochure hybrid only', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Cosmetic Sclerotherapy brochure', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Conservative treatment brochures', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Review us flyers', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Financial Hardship Form', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Pens vein only', '300', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Procedure Instructions Folders vein only', '300', 'Marketing'),
    ('Marketing', 'Bulk Order', 'No late legs insert', '300', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Pre and Post Instructions Eng Esp Vein', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Pre and Post Instructions Eng Esp Pain', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Pre and Post Instructions other RUS', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Vein Procedure form', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'EOB flyer', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Clinic Envelopes', '250', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Before and After Floor Decals', '# of rooms', 'Marketing'),
    ('Marketing', 'Bulk Order', 'Review badges for staff', '# of staff', 'Marketing'),
    # MARKETING - Laminated/Stick-On
    ('Marketing', 'Laminated/Stick-On', 'Late policy sign', '1', 'Marketing'),
    ('Marketing', 'Laminated/Stick-On', 'Back in 5 minutes sign', '1', 'Marketing'),
    ('Marketing', 'Laminated/Stick-On', 'Virtual FD sign', '1', 'Marketing'),
    ('Marketing', 'Laminated/Stick-On', 'Lunch Sign', '1', 'Marketing'),
    ('Marketing', 'Laminated/Stick-On', 'No Sun Sign', '1', 'Marketing'),
    ('Marketing', 'Laminated/Stick-On', 'Travel Plans Sign', '1', 'Marketing'),
]

CONSTRUCTION_TASKS_TEMPLATE = [
    'Construction Kick-Off Call',
    'Get Construction Timeline and Enter Dates Below',
    'Permitting',
    'Demolition',
    'Layout Walls',
    'Revise HVAC Ductwork',
    'Electrical Above Ceiling',
    'Framing',
    'MEP Rough-In',
    'One Side Sheetrock',
    'Rough Inspection',
    'Two Side Sheetrock',
    'Wall Inspection',
    'Tape/Finish Walls',
    'Paint',
    'Ceiling Grid Layout',
    'Above Ceiling Inspection',
    'Sprinkler Modifications',
    'Install/Repair Ceiling Tiles',
    'Millwork',
    'Install Doors and Hardware',
    'Install Flooring',
    'Final Inspections',
    'Final Clean',
    'Punch Walk',
]

# ─── Database ───────────────────────────────────────────────────────────────

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        db.execute("PRAGMA journal_mode=WAL")
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def query_db(query, args=(), one=False):
    cur = get_db().execute(query, args)
    rv = cur.fetchall()
    return (rv[0] if rv else None) if one else rv

def execute_db(query, args=()):
    db = get_db()
    cur = db.execute(query, args)
    db.commit()
    return cur.lastrowid

def init_db():
    db = sqlite3.connect(DATABASE)
    db.row_factory = sqlite3.Row
    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'team_member',
            department TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS clinics (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            opening_date DATE,
            status TEXT DEFAULT 'Active',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            is_template INTEGER DEFAULT 0,
            state TEXT,
            entity TEXT,
            sg_project_manager TEXT,
            sg_onsite_member TEXT,
            setup_week TEXT,
            ops_onsite_member TEXT,
            doctor_status TEXT,
            site_status TEXT,
            targeting_opening_month TEXT,
            procedures_done_where TEXT,
            paired_clinic TEXT,
            total_buildout_cost TEXT,
            cost_to_vip TEXT,
            data_analysis_summary TEXT,
            clinic_notes TEXT,
            photos_videos_link TEXT,
            lease_filename TEXT,
            supply_checkin_filename TEXT,
            lease_signed_date TEXT
        );
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clinic_id INTEGER,
            name TEXT NOT NULL,
            department TEXT NOT NULL,
            time_phase TEXT,
            due_date DATE,
            status TEXT DEFAULT 'Not Started',
            assignees TEXT DEFAULT '[]',
            order_index INTEGER DEFAULT 0,
            is_template INTEGER DEFAULT 0,
            template_offset_days INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (clinic_id) REFERENCES clinics(id)
        );
        CREATE TABLE IF NOT EXISTS notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            author_id INTEGER NOT NULL,
            content TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (task_id) REFERENCES tasks(id),
            FOREIGN KEY (author_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            uploaded_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (task_id) REFERENCES tasks(id)
        );
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            message TEXT NOT NULL,
            link TEXT,
            is_read INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clinic_id INTEGER,
            task_id INTEGER,
            user_id INTEGER,
            action TEXT NOT NULL,
            detail TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS lease_summaries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clinic_id INTEGER UNIQUE NOT NULL,
            landlord_name TEXT,
            landlord_contact TEXT,
            landlord_broker TEXT,
            lease_start_date DATE,
            lease_end_date DATE,
            monthly_rent TEXT,
            security_deposit TEXT,
            ti_allowance TEXT,
            rent_commencement_date DATE,
            square_footage TEXT,
            suite_unit TEXT,
            building_address TEXT,
            key_lease_terms TEXT,
            attorney_name TEXT,
            docusign_sent_to TEXT DEFAULT 'Matt Stearns - CFO',
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (clinic_id) REFERENCES clinics(id)
        );
        CREATE TABLE IF NOT EXISTS supply_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clinic_id INTEGER NOT NULL,
            tab TEXT NOT NULL,
            category TEXT NOT NULL,
            item_name TEXT NOT NULL,
            default_qty TEXT,
            amount_ordered TEXT,
            team_ordering TEXT,
            order_status TEXT DEFAULT 'Not Ordered',
            expected_delivery_date DATE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (clinic_id) REFERENCES clinics(id)
        );
        CREATE TABLE IF NOT EXISTS construction_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clinic_id INTEGER NOT NULL,
            name TEXT NOT NULL,
            start_date DATE,
            due_date DATE,
            status TEXT DEFAULT 'Not Started',
            notes TEXT,
            order_index INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (clinic_id) REFERENCES clinics(id)
        );
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clinic_id INTEGER NOT NULL REFERENCES clinics(id),
            expense_date TEXT,
            category TEXT NOT NULL,
            description TEXT,
            vendor TEXT,
            amount REAL DEFAULT 0,
            created_by INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS weekly_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clinic_id INTEGER NOT NULL REFERENCES clinics(id),
            week_ending TEXT NOT NULL,
            status TEXT DEFAULT 'On Track',
            highlights TEXT,
            next_week TEXT,
            timeline_changes TEXT,
            approvals_needed TEXT,
            help_needed TEXT,
            created_by INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(clinic_id, week_ending)
        );
        CREATE TABLE IF NOT EXISTS report_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            fields TEXT NOT NULL,
            created_by INTEGER REFERENCES users(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    db.commit()

    # Run migrations for existing databases
    new_clinic_cols = [
        ('state', 'TEXT'), ('entity', 'TEXT'), ('sg_project_manager', 'TEXT'),
        ('sg_onsite_member', 'TEXT'), ('setup_week', 'TEXT'), ('ops_onsite_member', 'TEXT'),
        ('doctor_status', 'TEXT'), ('site_status', 'TEXT'), ('targeting_opening_month', 'TEXT'),
        ('procedures_done_where', 'TEXT'), ('paired_clinic', 'TEXT'),
        ('total_buildout_cost', 'TEXT'), ('cost_to_vip', 'TEXT'),
        ('data_analysis_summary', 'TEXT'), ('clinic_notes', 'TEXT'),
        ('photos_videos_link', 'TEXT'), ('lease_filename', 'TEXT'),
        ('supply_checkin_filename', 'TEXT'),
        ('lease_signed_date', 'TEXT'),
        ('buildout_cost', 'REAL'), ('ti_allowance', 'REAL'), ('ti_paid_by_vip', 'REAL'),
    ]
    existing_cols = [row[1] for row in db.execute("PRAGMA table_info(clinics)").fetchall()]
    for col_name, col_type in new_clinic_cols:
        if col_name not in existing_cols:
            db.execute(f"ALTER TABLE clinics ADD COLUMN {col_name} {col_type}")

    # Migrate lease_summaries to add new rent fields
    new_lease_cols = [
        ('base_rent', 'TEXT'), ('cam_rent', 'TEXT'), ('total_monthly_rent', 'TEXT'),
        ('annual_rent_increase_pct', 'TEXT'), ('buildout_cost_estimate', 'TEXT'),
    ]
    existing_lease_cols = [row[1] for row in db.execute("PRAGMA table_info(lease_summaries)").fetchall()]
    for col_name, col_type in new_lease_cols:
        if col_name not in existing_lease_cols:
            db.execute(f"ALTER TABLE lease_summaries ADD COLUMN {col_name} {col_type}")
    db.commit()

    # Seed admin user
    admin = db.execute("SELECT id FROM users WHERE username='kelly'").fetchone()
    if not admin:
        db.execute(
            "INSERT INTO users (username, password_hash, full_name, role) VALUES (?,?,?,?)",
            ('kelly', generate_password_hash('VIPAdmin1!'), 'Kelly (Admin)', 'admin')
        )
        db.commit()

    # Seed template clinic
    tmpl = db.execute("SELECT id FROM clinics WHERE is_template=1").fetchone()
    if not tmpl:
        cur = db.execute(
            "INSERT INTO clinics (name, status, is_template) VALUES (?,?,?)",
            ('Master Template', 'Template', 1)
        )
        tmpl_id = cur.lastrowid
        db.commit()
        seed_template_tasks(db, tmpl_id)

    db.close()

def seed_template_tasks(db, clinic_id):
    tasks = []
    # Strategic Growth
    tasks += [
        ('General demographic research on possible locations', 'Strategic Growth', 'Scouting', -150, 0),
        ('Start scouting - rank and prioritize locations from demographic research', 'Strategic Growth', 'Scouting', -150, 1),
        ('Visit locations of interest', 'Strategic Growth', 'Scouting', -150, 2),
        ('Pull demographics specific to each location (during scouting)', 'Strategic Growth', 'Scouting', -150, 3),
        ('Get scaled floor plan from landlords (3rd scouting)', 'Strategic Growth', 'Scouting', -150, 4),
        ('Create marked up floor plan and work letter for landlords/contractors', 'Strategic Growth', 'After Lease Signed', -120, 5),
        ('Determine locations to send RFP (marked up floor plans)', 'Strategic Growth', 'After Lease Signed', -120, 6),
        ('Review proposals received back', 'Strategic Growth', 'After Lease Signed', -120, 7),
        ('Send counter proposals', 'Strategic Growth', 'After Lease Signed', -120, 8),
        ('Get estimate from contractors on build out cost / discuss TI allowance', 'Strategic Growth', 'After Lease Signed', -120, 9),
        ('Receive and review lease draft', 'Strategic Growth', 'After Lease Signed', -120, 10),
        ('Red line the lease draft', 'Strategic Growth', 'After Lease Signed', -120, 11),
        ('Send red lined lease to attorney for review', 'Strategic Growth', 'After Lease Signed', -120, 12),
        ('Submit red lined lease to landlord/broker for final lease production', 'Strategic Growth', 'After Lease Signed', -120, 13),
        ('Send email to Matt Stearns (CFO) with DocuSign timing, sender info, and key lease terms', 'Strategic Growth', 'After Lease Signed', -120, 14),
        ('File fully executed lease copy in lease folder', 'Strategic Growth', 'After Lease Signed', -120, 15),
        ('Send introductory questions to landlord/building manager after lease signed', 'Strategic Growth', 'After Lease Signed', -120, 16),
        ('Inform RCM about new location in credentialing meeting (after lease signed)', 'Strategic Growth', 'After Lease Signed', -120, 17),
        ('Send IT: floor plan, door lock mechanism pictures, site address, contact info', 'Strategic Growth', 'After Lease Signed', -120, 18),
        ('Send new location address to HR', 'Strategic Growth', 'After Lease Signed', -120, 19),
        ('Send marketing: signage info, measurements, pictures, recommended sign companies', 'Strategic Growth', 'After Lease Signed', -120, 20),
        ('Send new clinic email to Accounts Payable with lease attached (rent and security deposit amounts)', 'Strategic Growth', 'After Lease Signed', -120, 21),
        ('Confirm build out cost', 'Strategic Growth', 'After Lease Signed', -120, 22),
        ('Confirm floor plan is correct', 'Strategic Growth', 'After Lease Signed', -120, 23),
        ('Select finishes', 'Strategic Growth', 'After Lease Signed', -120, 24),
        ('Review anticipated timeline', 'Strategic Growth', 'After Lease Signed', -120, 25),
        ('Confirm correct contact info for build out', 'Strategic Growth', 'After Lease Signed', -120, 26),
        ('Send email confirming entire build out discussion', 'Strategic Growth', 'After Lease Signed', -120, 27),
        ('Send accounting final build out amounts', 'Strategic Growth', 'After Lease Signed', -120, 28),
        ('Notify AP to set up new Divvy card for purchases', 'Strategic Growth', 'After Lease Signed', -120, 29),
        ('Create liability insurance through LC', 'Strategic Growth', 'After Lease Signed', -120, 30),
        ('Determine PO box for checks', 'Strategic Growth', 'After Lease Signed', -120, 31),
        ('Confirm exam chair orders with Hill Beds', 'Strategic Growth', 'After Lease Signed', -120, 32),
        ('Send diplomas and medical licenses of doctors to marketing (request digital signature if new doctor)', 'Strategic Growth', 'After Lease Signed', -120, 33),
        ('Order ultrasound machines', 'Strategic Growth', '2 Months Before Opening', -60, 0),
        ('Send liability to general liability insurance (Tiffany G, Kelly Jackson)', 'Strategic Growth', '2 Months Before Opening', -60, 1),
        ('Review lease to understand all terms and services needed (cleaning, electricity, etc.)', 'Strategic Growth', '2 Months Before Opening', -60, 2),
        ('Confirm setup and first day plan (who sets up, who is there opening day)', 'Strategic Growth', '2 Months Before Opening', -60, 3),
        ('Send new clinic introduction and updates email', 'Strategic Growth', '2 Months Before Opening', -60, 4),
        ('Audit websites (coming soon page up, GMB created)', 'Strategic Growth', '2 Months Before Opening', -60, 5),
        ('Schedule fill tech for ultrasound setup', 'Strategic Growth', '1 Month Before Opening', -30, 0),
        ('Lease review (confirm everything clear)', 'Strategic Growth', '1 Month Before Opening', -30, 1),
        ('Confirm signage gets installed before first day', 'Strategic Growth', '1 Month Before Opening', -30, 2),
        ('Add clinic photo and opening date to new clinic opening email task', 'Strategic Growth', '1 Month Before Opening', -30, 3),
        ('Send opening email to staff at this clinic (photo + opening date)', 'Strategic Growth', '1 Month Before Opening', -30, 4),
        ('Add opening date to VIP calendar', 'Strategic Growth', '1 Month Before Opening', -30, 5),
        ('Create supply check-in sheet', 'Strategic Growth', '1 Month Before Opening', -30, 6),
        ('Confirm chair delivery time', 'Strategic Growth', '1 Month Before Opening', -30, 7),
        ('Audit websites and GMB for contact info', 'Strategic Growth', '1 Month Before Opening', -30, 8),
        ('Send construction updates email to staff', 'Strategic Growth', '1 Month Before Opening', -30, 9),
        ('Ask if any staff available to receive deliveries', 'Strategic Growth', '2 Weeks Before Opening', -14, 0),
        ('Make sure new door lock system is set up', 'Strategic Growth', '2 Weeks Before Opening', -14, 1),
        ('Send final construction updates', 'Strategic Growth', '2 Weeks Before Opening', -14, 2),
        ('Order credit card machine', 'Strategic Growth', '2 Weeks Before Opening', -14, 3),
        ('Schedule cleaners', 'Strategic Growth', '2 Weeks Before Opening', -14, 4),
        ('Hire handyman for furniture assembly/mounting', 'Strategic Growth', '2 Weeks Before Opening', -14, 5),
        ('Order furniture', 'Strategic Growth', '2 Weeks Before Opening', -14, 6),
        ('Make sure IT has door locks set up', 'Strategic Growth', '1 Week Before Opening', -7, 0),
        ('Make sure internet is set up', 'Strategic Growth', '1 Week Before Opening', -7, 1),
        ('Set up computers, iPads, ultrasounds', 'Strategic Growth', '1 Week Before Opening', -7, 2),
        ('Update location lease summary', 'Strategic Growth', 'Week Before Opening', -5, 0),
        ('Take clinic videos and written directions (for call center and Hub)', 'Strategic Growth', 'Week Before Opening', -5, 1),
        ('Verify scanners, printers, and label printer functioning', 'Strategic Growth', 'Week Before Opening', -5, 2),
        ('Test payment terminal', 'Strategic Growth', 'Week Before Opening', -5, 3),
        ('Confirm cleaning crew performance', 'Strategic Growth', 'Week Before Opening', -5, 4),
        ('Ensure emergency kit is set up', 'Strategic Growth', 'Week Before Opening', -5, 5),
        ('Confirm trash pickup schedule', 'Strategic Growth', 'Week Before Opening', -5, 6),
        ('Check in on deliveries', 'Strategic Growth', 'Week Before Opening', -5, 7),
        ('Verify exam rooms are stocked', 'Strategic Growth', 'Week Before Opening', -5, 8),
        ('Sunday check: sign on door, on directory, delivery sign up, monument sign up', 'Strategic Growth', 'Week Before Opening', -5, 9),
        ('Stock pantries/break room', 'Strategic Growth', 'Week Before Opening', -5, 10),
        ('Send "Actions Needed" email to ROD (day before opening)', 'Strategic Growth', 'Week Before Opening', -5, 11),
        ('Send "Important Information for Your First Day" email to all staff at new clinic (day before opening)', 'Strategic Growth', 'Week Before Opening', -5, 12),
        ('Final lease review - make sure everything in lease summary and staff informed', 'Strategic Growth', 'Week Before Opening', -5, 13),
        ('Send email to accounting with asset list (exam chairs, ultrasound, major assets now in use)', 'Strategic Growth', 'Opening Day', 0, 0),
        ('Send first week check-in email to ROD', 'Strategic Growth', '1 Week After Opening', 7, 0),
        ('Send first month check-in email to ROD', 'Strategic Growth', '1 Month After Opening', 30, 0),
    ]
    # Marketing
    tasks += [
        ('Prepare building signage', 'Marketing', 'After Lease Signed', -120, 0),
        ('Create print materials', 'Marketing', 'After Lease Signed', -120, 1),
        ('Coordinate photo shoot of new doctor with Carly', 'Marketing', 'After Lease Signed', -120, 2),
        ('Set up GMB, Yelp, Apple Maps, Bing pages', 'Marketing', 'After Lease Signed', -120, 3),
        ('Set up all websites and microsites', 'Marketing', 'After Lease Signed', -120, 4),
        ('Verify GMB page is set up correctly and working', 'Marketing', '2 Months Before Opening', -60, 0),
        ('Update marketing tab on supply check-in sheet (confirm quantities and supplies correct for this clinic)', 'Marketing', '2 Months Before Opening', -60, 1),
        ("Create ads (don't start them yet)", 'Marketing', '1 Month Before Opening', -30, 0),
        ('Add clinic information to website, microsites, and GMB', 'Marketing', '1 Month Before Opening', -30, 1),
        ('Order marketing supplies from Ops based on supply check-in sheet', 'Marketing', '1 Month Before Opening', -30, 2),
        ('Turn on ads', 'Marketing', '2 Weeks Before Opening', -14, 0),
        ('Make sure signs are installed', 'Marketing', '2 Weeks Before Opening', -14, 1),
        ('Send marketing materials feedback form', 'Marketing', '1 Month After Opening', 30, 0),
    ]
    # IT
    tasks += [
        ('Determine IT setup plan and communicate to Strategic Growth', 'IT', 'After Lease Signed', -120, 0),
        ('Schedule internet setup', 'IT', '2 Months Before Opening', -60, 0),
        ('Order tech items and arrange delivery', 'IT', '2 Months Before Opening', -60, 1),
        ('Confirm internet setup dates with Strategic Growth team', 'IT', '2 Months Before Opening', -60, 2),
        ('Create NextTech elements and clinic ops check', 'IT', '1 Month Before Opening', -30, 0),
        ('Link digital signature to NextTech', 'IT', '1 Month Before Opening', -30, 1),
        ('Generate provider profile in NextTech', 'IT', '1 Month Before Opening', -30, 2),
        ("Access provider's account and add signature as default signature (with date)", 'IT', '1 Month Before Opening', -30, 3),
        ('Change preferences to automatically change status of eMins signed to "Sign by Provider"', 'IT', '1 Month Before Opening', -30, 4),
        ('Add new location to NextTech (ASC or regular location)', 'IT', '1 Month Before Opening', -30, 5),
        ('EMR setup: add provider info to ultrasound note', 'IT', '1 Month Before Opening', -30, 6),
        ('Set up ultrasounds, TVs, iPads, Sonos, security cameras, and computers', 'IT', '1 Week Before Opening', -7, 0),
        ('Designate person to be on call for first day', 'IT', '1 Week Before Opening', -7, 1),
        ('Add location to facility/badge access system', 'IT', '1 Week Before Opening', -7, 2),
    ]
    # Inventory
    tasks += [
        ('Set up Medtronic account', 'Inventory', '1 Month Before Opening', -30, 0),
        ('Set up Besse account', 'Inventory', '1 Month Before Opening', -30, 1),
        ('Set up McKesson account', 'Inventory', '1 Month Before Opening', -30, 2),
        ('Set up Total Vein account', 'Inventory', '1 Month Before Opening', -30, 3),
        ('Set up Carolon Compression Stockings account', 'Inventory', '1 Month Before Opening', -30, 4),
        ('Set up Asclera Pine Pharmacy / Methapharm account', 'Inventory', '1 Month Before Opening', -30, 5),
        ('Set up Boston Scientific / BTG account', 'Inventory', '1 Month Before Opening', -30, 6),
        ('Set up Airgas account', 'Inventory', '1 Month Before Opening', -30, 7),
        ('Set up Stericycle account', 'Inventory', '1 Month Before Opening', -30, 8),
        ('Set up Water Cooler Company account', 'Inventory', '1 Month Before Opening', -30, 9),
        ('Order supplies from supply check-in sheet', 'Inventory', '1 Month Before Opening', -30, 10),
        ('Add expiration dates for emergency meds to system', 'Inventory', '1 Week Before Opening', -7, 0),
        ('Determine who will be doing inventory for this location', 'Inventory', '1 Week Before Opening', -7, 1),
        ('Finalize with Strategic Growth that all initial supplies were received and transitioning to formal inventory orders', 'Inventory', 'Week Before Opening', -5, 0),
        ('Order any additional items requested during opening week (items added to supply check-in sheet during opening)', 'Inventory', '1 Month After Opening', 30, 0),
    ]
    # Operations
    tasks += [
        ('Create emergency number for the doctors', 'Operations', 'After Lease Signed', -120, 0),
        ('Send New Location Alert to all teams (call center, scheduling, verifications, patient coordinators, RCM, etc.)', 'Operations', 'After Lease Signed', -120, 1),
        ("Update the Hub - doctor info up to date and clinic linked to doctor's site", 'Operations', 'After Lease Signed', -120, 2),
        ('Create site for new clinic on the Hub', 'Operations', 'After Lease Signed', -120, 3),
        ('Update schedules on Hub', 'Operations', 'After Lease Signed', -120, 4),
        ('Create Medwork location - NextTech mapping', 'Operations', 'After Lease Signed', -120, 5),
        ('Create Medwork location - Medwork 1.0 and 2.0', 'Operations', 'After Lease Signed', -120, 6),
        ('Create front desk email', 'Operations', 'After Lease Signed', -120, 7),
        ('Add location to ClearWave', 'Operations', '1 Month Before Opening', -30, 0),
        ('Add schedule to Hub', 'Operations', '1 Month Before Opening', -30, 1),
        ('Confirm schedule is open', 'Operations', '1 Month Before Opening', -30, 2),
        ('Add location and doctor to Luma', 'Operations', '1 Month Before Opening', -30, 3),
        ('Test check-in workflows with mock patient', 'Operations', '1 Week Before Opening', -7, 0),
        ('Determine key needs and make appropriate copies', 'Operations', '1 Week Before Opening', -7, 1),
        ('Determine opening and closing protocols', 'Operations', '1 Week After Opening', 7, 0),
        ('Run new clinic audit', 'Operations', '1 Month After Opening', 30, 0),
        ('One month check', 'Operations', '1 Month After Opening', 30, 1),
        ('Two month check', 'Operations', '2 Months After Opening', 60, 0),
        ('Three month check', 'Operations', '3 Months After Opening', 90, 0),
    ]
    # Accounting / Accounts Payable
    tasks += [
        ('Confirm rent payment is set up', 'Accounting / Accounts Payable', 'After Lease Signed', -120, 0),
        ('Confirm security deposit is sent out', 'Accounting / Accounts Payable', 'After Lease Signed', -120, 1),
        ('Add clinic to check management protocol ClickUp space (https://app.clickup.com/t/86dzk45rq)', 'Accounting / Accounts Payable', '1 Month Before Opening', -30, 0),
    ]
    # Credentialing
    tasks += [
        ('Send welcome email to new provider and start collecting credentialing documents', 'Credentialing', 'When New Provider Hired', 0, 0),
        ('Initiate out-of-network linking to PPO plans', 'Credentialing', 'After Lease Signed', -120, 0),
        ('Start credentialing enrollment for HMOs (if applicable, e.g. California)', 'Credentialing', 'After Lease Signed', -120, 1),
        ('Start Medicare credentialing enrollment for the provider', 'Credentialing', 'After Lease Signed', -120, 2),
        ('Confirm what plans the provider is fully linked with and participating in', 'Credentialing', '1 Week Before Opening', -7, 0),
        ('Share confirmed plan participation with Operations and Strategic Growth', 'Credentialing', '1 Week Before Opening', -7, 1),
    ]
    # RODs & Clinic Leads
    tasks += [
        ('Determine if this will be a virtual front desk location; inform Strategic Growth', 'RODs & Clinic Leads', 'After Lease Signed', -120, 0),
        ('Hire new staff and start new staff training (if not started beforehand)', 'RODs & Clinic Leads', 'After Lease Signed', -120, 1),
        ('Verify that front desk, office staff, and MAs can be hired and trained in time; confirm with Strategic Growth that opening date is feasible', 'RODs & Clinic Leads', 'After Lease Signed', -120, 2),
        ('Set up ultrasound training for staff', 'RODs & Clinic Leads', '1 Month Before Opening', -30, 0),
        ('Check out proper keys and key cards to each staff member; keep a list of who has what keys/key cards', 'RODs & Clinic Leads', 'Opening Day', 0, 0),
        ('Do Unified door training (how to arm/disarm alarm, set door open for certain hours, re-lock, use Unified door system)', 'RODs & Clinic Leads', 'Opening Day', 0, 1),
        ('Confirm iPads are logged into correct accounts and working', 'RODs & Clinic Leads', 'Opening Day', 0, 2),
        ('Confirm ultrasound uploads are working and images uploading to correct location in NextTech', 'RODs & Clinic Leads', 'Opening Day', 0, 3),
        ('Walk through waiting area — check cleanliness and seating before opening to patients', 'RODs & Clinic Leads', 'Opening Day', 0, 4),
        ('Ensure downtime packets are printed and accessible (in case internet goes down)', 'RODs & Clinic Leads', 'Opening Day', 0, 5),
        ('Observe patient check-in process and give feedback', 'RODs & Clinic Leads', 'Opening Day', 0, 6),
        ("Shadow provider's first consult — observe and give feedback", 'RODs & Clinic Leads', 'Opening Day', 0, 7),
        ('Confirm patient movement flow is smooth (waiting room → scan → provider → checkout)', 'RODs & Clinic Leads', 'Opening Day', 0, 8),
        ('Identify any bottlenecks or delays and address them', 'RODs & Clinic Leads', 'Opening Day', 0, 9),
        ('Provide breakfast for the team on the first day', 'RODs & Clinic Leads', 'Opening Day', 0, 10),
        ('Hold team huddle at start of day (front desk, sonographers, provider)', 'RODs & Clinic Leads', 'Opening Day', 0, 11),
        ('Review processes with front desk — ensure they can apply training and are ready', 'RODs & Clinic Leads', 'Opening Day', 0, 12),
        ('Print out Mercy Manual and put in red binder', 'RODs & Clinic Leads', 'Opening Day', 0, 13),
        ("Ensure everyone in clinic knows where Mercy Kit is located and what's in it", 'RODs & Clinic Leads', 'Opening Day', 0, 14),
        ('Get safe set up and go over safe protocol', 'RODs & Clinic Leads', 'Opening Day', 0, 15),
        ('Finish receiving and checking in all deliveries on supply check-in sheet', 'RODs & Clinic Leads', '1 Week After Opening', 7, 0),
        ("Identify any additional supplies needed that haven't been ordered; add to supply check-in sheet", 'RODs & Clinic Leads', '1 Week After Opening', 7, 1),
        ('Verify daily ultrasound images are uploading consistently to PACS', 'RODs & Clinic Leads', '1 Week After Opening', 7, 2),
        ('Audit image quality', 'RODs & Clinic Leads', '1 Week After Opening', 7, 3),
        ('Confirm cleaning staff schedule and that clinic team knows when they come in', 'RODs & Clinic Leads', '1 Week After Opening', 7, 4),
        ('Review supply inventory process, how to use inventory sheet, and how to reorder supplies with clinic team', 'RODs & Clinic Leads', '1 Week After Opening', 7, 5),
        ('Confirm consents are being completed properly in NextTech', 'RODs & Clinic Leads', '1 Week After Opening', 7, 6),
        ('Provide summary of first week observations to Strategic Growth (feedback on improvements)', 'RODs & Clinic Leads', '1 Week After Opening', 7, 7),
        ('Review patient feedback and wait times', 'RODs & Clinic Leads', '1 Month After Opening', 30, 0),
        ('Confirm techs are using correct measurements and protocols', 'RODs & Clinic Leads', '1 Month After Opening', 30, 1),
        ('Review abnormal studies with providers', 'RODs & Clinic Leads', '1 Month After Opening', 30, 2),
        ('Identify top 5 operational gaps', 'RODs & Clinic Leads', '1 Month After Opening', 30, 3),
        ('Review scheduling patterns (no-shows, bottlenecks)', 'RODs & Clinic Leads', '1 Month After Opening', 30, 4),
    ]
    # Malpractice
    tasks += [
        ('Create new malpractice insurance for the new provider with the correct new location (if new provider)', 'Malpractice', 'After Lease Signed', -120, 0),
        ("Add the new location to the provider's COI - Certificate of Insurance (if existing provider)", 'Malpractice', 'After Lease Signed', -120, 1),
    ]
    # Special Operations
    tasks += [
        ('Update ClickUp MR mapping', 'Special Operations', 'After Lease Signed', -120, 0),
        ('Update ClickUp Insurance Billing mapping', 'Special Operations', 'After Lease Signed', -120, 1),
        ('Update Saturation Report', 'Special Operations', 'After Lease Signed', -120, 2),
        ('Update Master Insurance Location', 'Special Operations', 'After Lease Signed', -120, 3),
        ('Update High Level Report', 'Special Operations', 'After Lease Signed', -120, 4),
        ('Update Collections Report', 'Special Operations', 'After Lease Signed', -120, 5),
        ('Update Bill Scrub Report', 'Special Operations', 'After Lease Signed', -120, 6),
        ('Update Discrepancy Report', 'Special Operations', 'After Lease Signed', -120, 7),
        ('Update Appointment Audit', 'Special Operations', 'After Lease Signed', -120, 8),
        ('Update Conservative Report', 'Special Operations', 'After Lease Signed', -120, 9),
        ('Update Reconciliation Report', 'Special Operations', 'After Lease Signed', -120, 10),
        ('Update Productivity mapping', 'Special Operations', 'After Lease Signed', -120, 11),
        ('Set up data reports', 'Special Operations', '1 Month Before Opening', -30, 0),
        ('Ramp up update', 'Special Operations', '1 Month Before Opening', -30, 1),
        ('Opening date setup', 'Special Operations', '1 Month Before Opening', -30, 2),
        ('Marketing data set update', 'Special Operations', '1 Month Before Opening', -30, 3),
        ('Amazon connection / Q configuration', 'Special Operations', '1 Month Before Opening', -30, 4),
        ('Google My Business (GMB) update', 'Special Operations', '1 Month Before Opening', -30, 5),
        ('ClickUp mapping', 'Special Operations', '1 Month Before Opening', -30, 6),
        ('NextTech mapping', 'Special Operations', '1 Month Before Opening', -30, 7),
        ('MPS update', 'Special Operations', '1 Month Before Opening', -30, 8),
    ]
    # Referrals Outreach
    tasks += [
        ('Begin referral outreach in areas being considered for new location', 'Referrals Outreach', 'Scouting', -150, 0),
        ('Build lead sourcing plan', 'Referrals Outreach', '2 Months Before Opening', -60, 0),
        ('Finalize top lead short list', 'Referrals Outreach', '2 Months Before Opening', -60, 1),
        ('Launch outreach', 'Referrals Outreach', '1 Month Before Opening', -30, 0),
        ('Activation check: establish partnerships + first referral watch + 45 days', 'Referrals Outreach', '1 Month Before Opening', -30, 1),
        ('Share short list with Regional Operations Directors for site visit', 'Referrals Outreach', '1 Month Before Opening', -30, 2),
        ('Send opening announcement touch to top targets (logged in HubSpot)', 'Referrals Outreach', '1 Week Before Opening', -7, 0),
        ('Order opening day marketing materials for referrals', 'Referrals Outreach', '1 Week Before Opening', -7, 1),
    ]
    # HR
    tasks += [
        ('Add new location to Rippling', 'HR', 'After Lease Signed', -120, 0),
        ('Order HR compliance posters (to be delivered the week before opening)', 'HR', '1 Month Before Opening', -30, 0),
        ('Add any new staff to this location in Rippling', 'HR', 'When New Provider Hired', 0, 0),
    ]

    for (name, dept, phase, offset, order) in tasks:
        db.execute(
            '''INSERT INTO tasks (clinic_id, name, department, time_phase, status, order_index,
               is_template, template_offset_days) VALUES (?,?,?,?,?,?,?,?)''',
            (clinic_id, name, dept, phase, 'Not Started', order, 1, offset)
        )
    db.commit()

def populate_supply_items(clinic_id):
    """Auto-populate supply items for a new clinic from master list."""
    for (tab, category, item_name, default_qty, team) in SUPPLY_MASTER:
        execute_db(
            '''INSERT INTO supply_items (clinic_id, tab, category, item_name, default_qty, team_ordering, order_status)
               VALUES (?,?,?,?,?,?,?)''',
            (clinic_id, tab, category, item_name, default_qty, team, 'Not Ordered')
        )

def populate_construction_tasks(clinic_id):
    """Auto-populate construction tasks for a new clinic."""
    for i, name in enumerate(CONSTRUCTION_TASKS_TEMPLATE):
        execute_db(
            '''INSERT INTO construction_tasks (clinic_id, name, status, order_index)
               VALUES (?,?,?,?)''',
            (clinic_id, name, 'Not Started', i)
        )

EXPENSE_TEMPLATES = [
    ('Exam Chairs', 'Down payment for exam chairs ($1488.70 x # of chairs)', 'Hill'),
    ('Exam Chairs', 'Final payment for exam chairs (see invoice)', 'Hill'),
    ('Ultrasounds', 'Ultrasound machines', 'GE'),
    ('Food', 'Breakfast for staff on first day', ''),
    ('Handyman', 'Taskrabbits for furniture building and mounting', 'Taskrabbit'),
    ('Furniture', 'Furniture for office', ''),
    ('Office Supplies', 'Office supplies', 'Amazon'),
    ('Medical Supplies', 'Non-ablation/varithena medical supplies', 'McKesson/Carolon/Total Vein/Asclera/Amazon'),
    ('Medical Supplies', 'Ablation materials (catheters/intros/starters)', 'Medtronic'),
    ('Medical Supplies', 'Varithena materials (canisters/admin packs)', 'Boston Scientific'),
    ('IT', 'IT Supplies (TVs/Sonos/Cameras/Mounts/Network Equipment/Door Lock)', 'Amazon/IT Direct'),
    ('IT', 'IT equipment installation and set-up', 'AllMedia'),
    ('Transportation', 'Ubers and transportation during setup', ''),
    ('Hotel', 'Hotel where PM stayed during set up', ''),
    ('Food', 'Food PM ate while at set up', ''),
    ('Flights', 'Flight to location for PM', ''),
]

def populate_expense_templates(clinic_id, user_id=None):
    """Auto-populate expense template rows for a new clinic."""
    for (category, description, vendor) in EXPENSE_TEMPLATES:
        execute_db(
            '''INSERT INTO expenses (clinic_id, category, description, vendor, amount, created_by)
               VALUES (?,?,?,?,0,?)''',
            (clinic_id, category, description, vendor, user_id)
        )

# ─── Auth helpers ────────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = query_db("SELECT * FROM users WHERE id=?", [session['user_id']], one=True)
        if not user or user['role'] != 'admin':
            flash('Admin access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def pm_required(f):
    """Allows admin and project_manager roles."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        user = query_db("SELECT * FROM users WHERE id=?", [session['user_id']], one=True)
        if not user or user['role'] not in ('admin', 'project_manager'):
            flash('Project Manager or Admin access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def current_user():
    if 'user_id' in session:
        return query_db("SELECT * FROM users WHERE id=?", [session['user_id']], one=True)
    return None

def unread_notification_count():
    if 'user_id' not in session:
        return 0
    row = query_db("SELECT COUNT(*) as cnt FROM notifications WHERE user_id=? AND is_read=0",
                   [session['user_id']], one=True)
    return row['cnt'] if row else 0

app.jinja_env.globals['current_user'] = current_user
app.jinja_env.globals['unread_count'] = unread_notification_count
app.jinja_env.filters['from_json'] = json.loads

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_doc_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_DOC_EXTENSIONS

def notify_user(user_id, message, link=None):
    execute_db("INSERT INTO notifications (user_id, message, link) VALUES (?,?,?)",
               (user_id, message, link))

def log_activity(clinic_id, task_id, user_id, action, detail=None):
    execute_db(
        "INSERT INTO activity_log (clinic_id, task_id, user_id, action, detail) VALUES (?,?,?,?,?)",
        (clinic_id, task_id, user_id, action, detail)
    )

def can_edit_task(user, task):
    if user['role'] in ('admin', 'project_manager'):
        return True
    if user['role'] == 'dept_head' and user['department'] == task['department']:
        return True
    if user['role'] == 'team_member':
        assignees = json.loads(task['assignees'] or '[]')
        if str(user['id']) in [str(a) for a in assignees]:
            return True
    return False

def get_quarter(date_str):
    if not date_str:
        return None
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        q = (dt.month - 1) // 3 + 1
        return f'Q{q} {dt.year}'
    except Exception:
        return None

def save_clinic_document(file_obj, clinic_id, doc_type):
    if not file_obj or not file_obj.filename:
        return None
    if not allowed_doc_file(file_obj.filename):
        return None
    upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'clinics', str(clinic_id))
    os.makedirs(upload_dir, exist_ok=True)
    filename = secure_filename(file_obj.filename)
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    saved_name = f'{doc_type}_{timestamp}_{filename}'
    file_obj.save(os.path.join(upload_dir, saved_name))
    return saved_name

def extract_lease_data(pdf_path):
    """Extract lease fields from PDF using PyMuPDF."""
    data = {}
    try:
        import fitz
        doc = fitz.open(pdf_path)
        text = ''
        for page in doc:
            text += page.get_text()
        doc.close()

        # Monthly rent
        rent_m = re.search(r'(?:monthly\s+(?:base\s+)?rent|base\s+rent)[^\$\d]*[\$]?\s*([\d,]+(?:\.\d{2})?)', text, re.IGNORECASE)
        if rent_m:
            data['monthly_rent'] = rent_m.group(1).replace(',', '')

        # Security deposit
        dep_m = re.search(r'security\s+deposit[^\$\d]*[\$]?\s*([\d,]+(?:\.\d{2})?)', text, re.IGNORECASE)
        if dep_m:
            data['security_deposit'] = dep_m.group(1).replace(',', '')

        # TI allowance
        ti_m = re.search(r'(?:tenant\s+improvement|TI)\s+allowance[^\$\d]*[\$]?\s*([\d,]+(?:\.\d{2})?)', text, re.IGNORECASE)
        if ti_m:
            data['ti_allowance'] = ti_m.group(1).replace(',', '')

        # Square footage
        sqft_m = re.search(r'([\d,]+)\s*(?:square\s+feet|sq\.?\s*ft\.?|rentable\s+(?:square\s+)?(?:feet|area))', text, re.IGNORECASE)
        if sqft_m:
            data['square_footage'] = sqft_m.group(1).replace(',', '')

        # Dates - lease start
        start_m = re.search(r'(?:lease\s+(?:commencement|start)\s+date|commencement\s+date)[^\d]*(\w+\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})', text, re.IGNORECASE)
        if start_m:
            data['lease_start_date'] = start_m.group(1)

        # Lease end
        end_m = re.search(r'(?:lease\s+(?:expiration|end|termination)\s+date|expiration\s+date)[^\d]*(\w+\s+\d{1,2},?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4})', text, re.IGNORECASE)
        if end_m:
            data['lease_end_date'] = end_m.group(1)

        # Landlord name
        ll_m = re.search(r'(?:landlord|lessor)[:\s]+([A-Z][A-Za-z\s,\.]+?)(?:\n|,|\.|LLC|Inc|Corp|LP)', text)
        if ll_m:
            data['landlord_name'] = ll_m.group(1).strip()

        # Address
        addr_m = re.search(r'(?:premises|property|located at)[:\s]+([0-9]+[^\n,]{5,60})', text, re.IGNORECASE)
        if addr_m:
            data['building_address'] = addr_m.group(1).strip()

    except Exception as e:
        pass
    return data

# ─── Routes ──────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = query_db("SELECT * FROM users WHERE username=?", [username], one=True)
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['full_name'] = user['full_name']
            return redirect(url_for('index'))
        flash('Invalid username or password.', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    clinics = query_db("SELECT * FROM clinics WHERE is_template=0 ORDER BY created_at DESC")
    clinic_stats = []
    for clinic in clinics:
        total = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=?", [clinic['id']], one=True)['cnt']
        done = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND status='Complete'", [clinic['id']], one=True)['cnt']
        blocked = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND status='Blocked'", [clinic['id']], one=True)['cnt']
        overdue = query_db(
            "SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND due_date < date('now') AND status NOT IN ('Complete')",
            [clinic['id']], one=True)['cnt']
        pct = round((done / total * 100) if total > 0 else 0)
        clinic_stats.append({'clinic': dict(clinic), 'total': total, 'done': done, 'pct': pct, 'blocked': blocked, 'overdue': overdue})
    return render_template('index.html', clinic_stats=clinic_stats)

@app.route('/clinic/new', methods=['GET', 'POST'])
@pm_required
def new_clinic():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        opening_date = request.form.get('opening_date', '') or None
        state = request.form.get('state', '') or None
        lease_signed_date = request.form.get('lease_signed_date', '' ) or None
        entity = request.form.get('entity', '') or None
        sg_pm = request.form.get('sg_project_manager', '') or None
        sg_onsite = request.form.get('sg_onsite_member', '') or None
        setup_week = request.form.get('setup_week', '') or None
        ops_onsite = request.form.get('ops_onsite_member', '') or None
        doctor_status = request.form.get('doctor_status', '') or None
        site_status = request.form.get('site_status', '') or None
        targeting_month = request.form.get('targeting_opening_month', '') or None
        procedures_done_where = request.form.get('procedures_done_where', '') or None
        paired_clinic = request.form.get('paired_clinic', '') or None
        total_buildout_cost = request.form.get('total_buildout_cost', '') or None
        cost_to_vip = request.form.get('cost_to_vip', '') or None
        data_analysis = request.form.get('data_analysis_summary', '') or None
        clinic_notes = request.form.get('clinic_notes', '') or None
        photos_link = request.form.get('photos_videos_link', '') or None

        if not name:
            flash('Clinic name is required.', 'error')
            return render_template('new_clinic.html')

        clinic_id = execute_db(
            '''INSERT INTO clinics (name, opening_date, created_by, state, entity,
               sg_project_manager, sg_onsite_member, setup_week, ops_onsite_member,
               doctor_status, site_status, targeting_opening_month, procedures_done_where,
               paired_clinic, total_buildout_cost, cost_to_vip, data_analysis_summary,
               clinic_notes, photos_videos_link)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (name, opening_date, session['user_id'], state, entity,
             sg_pm, sg_onsite, setup_week, ops_onsite,
             doctor_status, site_status, targeting_month, procedures_done_where,
             paired_clinic, total_buildout_cost, cost_to_vip, data_analysis,
             clinic_notes, photos_link)
        )

        # Handle file uploads
        lease_file = request.files.get('lease_file')
        supply_file = request.files.get('supply_checkin_file')
        if lease_file and lease_file.filename:
            saved = save_clinic_document(lease_file, clinic_id, 'lease')
            if saved:
                execute_db("UPDATE clinics SET lease_filename=? WHERE id=?", (saved, clinic_id))
        if supply_file and supply_file.filename:
            saved = save_clinic_document(supply_file, clinic_id, 'supply')
            if saved:
                execute_db("UPDATE clinics SET supply_checkin_filename=? WHERE id=?", (saved, clinic_id))

        # Copy template tasks
        tmpl = query_db("SELECT * FROM clinics WHERE is_template=1", one=True)
        if tmpl:
            tmpl_tasks = query_db("SELECT * FROM tasks WHERE clinic_id=?", [tmpl['id']])
            for t in tmpl_tasks:
                due = None
                if opening_date and t['template_offset_days'] is not None:
                    od = datetime.strptime(opening_date, '%Y-%m-%d')
                    due = (od + timedelta(days=t['template_offset_days'])).strftime('%Y-%m-%d')
                execute_db(
                    '''INSERT INTO tasks (clinic_id, name, department, time_phase, due_date, status,
                       order_index, template_offset_days) VALUES (?,?,?,?,?,?,?,?)''',
                    (clinic_id, t['name'], t['department'], t['time_phase'], due,
                     'Not Started', t['order_index'], t['template_offset_days'])
                )

        # Auto-populate supplies, construction tasks, and expense templates
        populate_supply_items(clinic_id)
        populate_construction_tasks(clinic_id)
        populate_expense_templates(clinic_id, session['user_id'])

        log_activity(clinic_id, None, session['user_id'], 'Clinic Created', name)
        flash(f'Clinic "{name}" created successfully!', 'success')
        return redirect(url_for('clinic_dashboard', clinic_id=clinic_id))
    return render_template('new_clinic.html')

@app.route('/clinic/<int:clinic_id>/edit', methods=['GET', 'POST'])
@pm_required
def edit_clinic(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    if request.method == 'POST':
        name = request.form.get('name', '').strip() or clinic['name']
        opening_date = request.form.get('opening_date', '') or None
        state = request.form.get('state', '') or None
        entity = request.form.get('entity', '') or None
        lease_signed_date = request.form.get('lease_signed_date', '' ) or None
        sg_pm = request.form.get('sg_project_manager', '') or None
        sg_onsite = request.form.get('sg_onsite_member', '') or None
        setup_week = request.form.get('setup_week', '') or None
        ops_onsite = request.form.get('ops_onsite_member', '') or None
        doctor_status = request.form.get('doctor_status', '') or None
        site_status = request.form.get('site_status', '') or None
        targeting_month = request.form.get('targeting_opening_month', '') or None
        procedures_done_where = request.form.get('procedures_done_where', '') or None
        paired_clinic = request.form.get('paired_clinic', '') or None
        total_buildout_cost = request.form.get('total_buildout_cost', '') or None
        cost_to_vip = request.form.get('cost_to_vip', '') or None
        data_analysis = request.form.get('data_analysis_summary', '') or None
        clinic_notes = request.form.get('clinic_notes', '') or None
        photos_link = request.form.get('photos_videos_link', '') or None

        execute_db(
            '''UPDATE clinics SET name=?, opening_date=?, state=?, entity=?,
               sg_project_manager=?, sg_onsite_member=?, setup_week=?, ops_onsite_member=?,
               doctor_status=?, site_status=?, targeting_opening_month=?,
               procedures_done_where=?, paired_clinic=?, total_buildout_cost=?, cost_to_vip=?,
               data_analysis_summary=?, clinic_notes=?, photos_videos_link=?
               WHERE id=?''',
            (name, opening_date, state, entity,
             sg_pm, sg_onsite, setup_week, ops_onsite,
             doctor_status, site_status, targeting_month,
             procedures_done_where, paired_clinic, total_buildout_cost, cost_to_vip,
             data_analysis, clinic_notes, photos_link, clinic_id)
        )

        lease_file = request.files.get('lease_file')
        supply_file = request.files.get('supply_checkin_file')
        if lease_file and lease_file.filename:
            saved = save_clinic_document(lease_file, clinic_id, 'lease')
            if saved:
                execute_db("UPDATE clinics SET lease_filename=? WHERE id=?", (saved, clinic_id))
        if supply_file and supply_file.filename:
            saved = save_clinic_document(supply_file, clinic_id, 'supply')
            if saved:
                execute_db("UPDATE clinics SET supply_checkin_filename=? WHERE id=?", (saved, clinic_id))

        log_activity(clinic_id, None, session['user_id'], 'Clinic Updated', name)
        flash(f'Clinic "{name}" updated successfully!', 'success')
        return redirect(url_for('clinic_dashboard', clinic_id=clinic_id))
    return render_template('edit_clinic.html', clinic=clinic)

@app.route('/clinic/<int:clinic_id>')
@login_required
def clinic_dashboard(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    tasks = query_db("SELECT * FROM tasks WHERE clinic_id=? ORDER BY department, template_offset_days, order_index", [clinic_id])
    dept_stats = {}
    for dept in DEPARTMENTS:
        dept_tasks = [t for t in tasks if t['department'] == dept]
        total = len(dept_tasks)
        done = sum(1 for t in dept_tasks if t['status'] == 'Complete')
        pct = round((done / total * 100) if total > 0 else 0)
        dept_stats[dept] = {'total': total, 'done': done, 'pct': pct}

    total_all = len(tasks)
    done_all = sum(1 for t in tasks if t['status'] == 'Complete')
    pct_all = round((done_all / total_all * 100) if total_all > 0 else 0)

    today = datetime.now().date()
    overdue = [t for t in tasks if t['due_date'] and datetime.strptime(t['due_date'], '%Y-%m-%d').date() < today and t['status'] not in ('Complete',)]
    blocked = [t for t in tasks if t['status'] == 'Blocked']

    activity = query_db(
        '''SELECT al.*, u.full_name, t.name as task_name FROM activity_log al
           LEFT JOIN users u ON al.user_id=u.id
           LEFT JOIN tasks t ON al.task_id=t.id
           WHERE al.clinic_id=? ORDER BY al.created_at DESC LIMIT 20''',
        [clinic_id]
    )
    return render_template('clinic_dashboard.html',
        clinic=clinic, tasks=tasks, dept_stats=dept_stats,
        pct_all=pct_all, overdue=overdue, blocked=blocked,
        activity=activity, departments=DEPARTMENTS, today=str(today))

@app.route('/clinic/<int:clinic_id>/tasks')
@login_required
def clinic_tasks(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        return redirect(url_for('index'))
    dept_filter = request.args.get('dept', '')
    phase_filter = request.args.get('phase', '')
    status_filter = request.args.get('status', '')
    q = "SELECT * FROM tasks WHERE clinic_id=?"
    args = [clinic_id]
    if dept_filter:
        q += " AND department=?"; args.append(dept_filter)
    if phase_filter:
        q += " AND time_phase=?"; args.append(phase_filter)
    if status_filter:
        q += " AND status=?"; args.append(status_filter)
    q += " ORDER BY department, template_offset_days, order_index"
    tasks = query_db(q, args)
    users = query_db("SELECT id, full_name, department FROM users ORDER BY full_name")
    today_str = str(datetime.now().date())
    return render_template('tasks.html', clinic=clinic, tasks=tasks, users=users,
                           departments=DEPARTMENTS, time_phases=TIME_PHASES, statuses=STATUSES,
                           dept_filter=dept_filter, phase_filter=phase_filter, status_filter=status_filter,
                           today_str=today_str)

@app.route('/clinic/<int:clinic_id>/task/new', methods=['GET', 'POST'])
@login_required
def new_task(clinic_id):
    user = current_user()
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        return redirect(url_for('index'))
    if user['role'] not in ('admin', 'project_manager', 'dept_head'):
        flash('Permission denied.', 'error')
        return redirect(url_for('clinic_tasks', clinic_id=clinic_id))
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        dept = request.form.get('department', '')
        phase = request.form.get('time_phase', '')
        due = request.form.get('due_date', '') or None
        status = request.form.get('status', 'Not Started')
        assignees = json.dumps(request.form.getlist('assignees'))
        execute_db(
            '''INSERT INTO tasks (clinic_id, name, department, time_phase, due_date, status, assignees)
               VALUES (?,?,?,?,?,?,?)''',
            (clinic_id, name, dept, phase, due, status, assignees)
        )
        log_activity(clinic_id, None, session['user_id'], 'Task Created', name)
        flash('Task created.', 'success')
        return redirect(url_for('clinic_tasks', clinic_id=clinic_id))
    users = query_db("SELECT id, full_name FROM users ORDER BY full_name")
    return render_template('task_form.html', clinic=clinic, task=None, users=users,
                           departments=DEPARTMENTS, time_phases=TIME_PHASES, statuses=STATUSES)

@app.route('/task/<int:task_id>', methods=['GET', 'POST'])
@login_required
def task_detail(task_id):
    task = query_db("SELECT * FROM tasks WHERE id=?", [task_id], one=True)
    if not task:
        flash('Task not found.', 'error')
        return redirect(url_for('index'))
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [task['clinic_id']], one=True)
    user = current_user()

    users = query_db("SELECT id, full_name, department FROM users ORDER BY full_name")
    notes = query_db(
        '''SELECT n.*, u.full_name FROM notes n JOIN users u ON n.author_id=u.id
           WHERE n.task_id=? ORDER BY n.created_at''', [task_id])
    attachments = query_db(
        '''SELECT a.*, u.full_name FROM attachments a JOIN users u ON a.uploaded_by=u.id
           WHERE a.task_id=? ORDER BY a.created_at''', [task_id])

    if request.method == 'POST':
        action = request.form.get('action')
        if not can_edit_task(user, task):
            flash('You do not have permission to update this task.', 'error')
            return redirect(url_for('task_detail', task_id=task_id))

        if action == 'update':
            name = request.form.get('name', task['name'])
            dept = request.form.get('department', task['department'])
            phase = request.form.get('time_phase', task['time_phase'])
            due = request.form.get('due_date', '') or None
            status = request.form.get('status', task['status'])
            assignees = json.dumps(request.form.getlist('assignees'))
            old_status = task['status']
            execute_db(
                '''UPDATE tasks SET name=?, department=?, time_phase=?, due_date=?, status=?,
                   assignees=?, updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                (name, dept, phase, due, status, assignees, task_id)
            )
            if old_status != status:
                log_activity(task['clinic_id'], task_id, session['user_id'],
                             'Status Changed', f'{old_status} → {status}')
                for uid in json.loads(assignees):
                    if int(uid) != session['user_id']:
                        notify_user(int(uid), f'Task "{name}" status changed to {status}',
                                   f'/task/{task_id}')
            flash('Task updated.', 'success')
            return redirect(url_for('task_detail', task_id=task_id))

        elif action == 'add_note':
            content = request.form.get('content', '').strip()
            if content:
                execute_db("INSERT INTO notes (task_id, author_id, content) VALUES (?,?,?)",
                           (task_id, session['user_id'], content))
                log_activity(task['clinic_id'], task_id, session['user_id'],
                             'Note Added', content[:80])
                mentions = re.findall(r'@(\w+)', content)
                for username in mentions:
                    mentioned = query_db("SELECT id FROM users WHERE username=?", [username], one=True)
                    if mentioned and mentioned['id'] != session['user_id']:
                        notify_user(mentioned['id'],
                                   f'{user["full_name"]} mentioned you in task "{task["name"]}"',
                                   f'/task/{task_id}')
                dept_users = query_db(
                    "SELECT id FROM users WHERE department=? AND id != ?",
                    [task['department'], session['user_id']])
                for du in dept_users:
                    notify_user(du['id'],
                               f'New note on task "{task["name"]}" in {task["department"]}',
                               f'/task/{task_id}')
                flash('Note added.', 'success')
            return redirect(url_for('task_detail', task_id=task_id))

        elif action == 'upload':
            if 'file' in request.files:
                f = request.files['file']
                if f and f.filename and allowed_file(f.filename):
                    filename = secure_filename(f.filename)
                    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                    saved_name = f'{timestamp}_{filename}'
                    f.save(os.path.join(app.config['UPLOAD_FOLDER'], saved_name))
                    execute_db(
                        "INSERT INTO attachments (task_id, filename, original_name, uploaded_by) VALUES (?,?,?,?)",
                        (task_id, saved_name, filename, session['user_id'])
                    )
                    log_activity(task['clinic_id'], task_id, session['user_id'], 'File Uploaded', filename)
                    flash('File uploaded.', 'success')
                else:
                    flash('Invalid file type.', 'error')
            return redirect(url_for('task_detail', task_id=task_id))

    assignee_ids = json.loads(task['assignees'] or '[]')
    return render_template('task_detail.html', task=task, clinic=clinic, user=user,
                           users=users, notes=notes, attachments=attachments,
                           assignee_ids=[str(a) for a in assignee_ids],
                           departments=DEPARTMENTS, time_phases=TIME_PHASES, statuses=STATUSES,
                           can_edit=can_edit_task(user, task))

@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/clinic-uploads/<int:clinic_id>/<filename>')
@login_required
def clinic_uploaded_file(clinic_id, filename):
    upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'clinics', str(clinic_id))
    return send_from_directory(upload_dir, filename)

# ─── Lease Summary ────────────────────────────────────────────────────────────

@app.route('/clinic/<int:clinic_id>/lease-summary', methods=['GET', 'POST'])
@login_required
def lease_summary(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    lease = query_db("SELECT * FROM lease_summaries WHERE clinic_id=?", [clinic_id], one=True)

    pdf_extracted = None
    if request.method == 'POST':
        action = request.form.get('action', 'save')

        # Handle PDF upload & extraction
        if action == 'upload_pdf':
            pdf_file = request.files.get('lease_pdf')
            if pdf_file and pdf_file.filename and pdf_file.filename.lower().endswith('.pdf'):
                saved = save_clinic_document(pdf_file, clinic_id, 'lease')
                if saved:
                    execute_db("UPDATE clinics SET lease_filename=? WHERE id=?", (saved, clinic_id))
                    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'clinics', str(clinic_id), saved)
                    pdf_extracted = extract_lease_data(pdf_path)
                    flash('PDF uploaded and data extracted. Review and save below.', 'success')
                    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
            return render_template('lease_summary.html', clinic=clinic, lease=lease, pdf_extracted=pdf_extracted)

        fields = {
            'landlord_name': request.form.get('landlord_name', '') or None,
            'landlord_contact': request.form.get('landlord_contact', '') or None,
            'landlord_broker': request.form.get('landlord_broker', '') or None,
            'lease_start_date': request.form.get('lease_start_date', '') or None,
            'lease_end_date': request.form.get('lease_end_date', '') or None,
            'base_rent': request.form.get('base_rent', '') or None,
            'cam_rent': request.form.get('cam_rent', '') or None,
            'total_monthly_rent': request.form.get('total_monthly_rent', '') or None,
            'annual_rent_increase_pct': request.form.get('annual_rent_increase_pct', '') or None,
            'monthly_rent': request.form.get('monthly_rent', '') or None,
            'security_deposit': request.form.get('security_deposit', '') or None,
            'ti_allowance': request.form.get('ti_allowance', '') or None,
            'buildout_cost_estimate': request.form.get('buildout_cost_estimate', '') or None,
            'rent_commencement_date': request.form.get('rent_commencement_date', '') or None,
            'square_footage': request.form.get('square_footage', '') or None,
            'suite_unit': request.form.get('suite_unit', '') or None,
            'building_address': request.form.get('building_address', '') or None,
            'key_lease_terms': request.form.get('key_lease_terms', '') or None,
            'attorney_name': request.form.get('attorney_name', '') or None,
            'docusign_sent_to': request.form.get('docusign_sent_to', 'Matt Stearns - CFO') or 'Matt Stearns - CFO',
        }
        if lease:
            sets = ', '.join(f'{k}=?' for k in fields.keys())
            execute_db(f"UPDATE lease_summaries SET {sets}, updated_at=CURRENT_TIMESTAMP WHERE clinic_id=?",
                       list(fields.values()) + [clinic_id])
        else:
            cols = ', '.join(fields.keys())
            placeholders = ', '.join('?' for _ in fields)
            execute_db(f"INSERT INTO lease_summaries (clinic_id, {cols}) VALUES (?, {placeholders})",
                       [clinic_id] + list(fields.values()))
        flash('Lease summary saved.', 'success')
        return redirect(url_for('lease_summary', clinic_id=clinic_id))

    lease = query_db("SELECT * FROM lease_summaries WHERE clinic_id=?", [clinic_id], one=True)
    return render_template('lease_summary.html', clinic=clinic, lease=lease, pdf_extracted=pdf_extracted)

# ─── Supply Check-In ──────────────────────────────────────────────────────────

@app.route('/clinic/<int:clinic_id>/supplies')
@login_required
def clinic_supplies(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    items = query_db("SELECT * FROM supply_items WHERE clinic_id=? ORDER BY tab, category, id", [clinic_id])
    # If no items exist yet, populate them
    if not items:
        populate_supply_items(clinic_id)
        items = query_db("SELECT * FROM supply_items WHERE clinic_id=? ORDER BY tab, category, id", [clinic_id])

    # Organize by tab
    tabs = ['Office', 'Medical', 'IT', 'Marketing', 'Post-Opening Orders', 'Returns']
    tab_data = {}
    for tab in tabs:
        tab_items = [i for i in items if i['tab'] == tab]
        # Group by category
        cats = {}
        for item in tab_items:
            cat = item['category']
            if cat not in cats:
                cats[cat] = []
            cats[cat].append(dict(item))
        # Progress
        total = len(tab_items)
        delivered = sum(1 for i in tab_items if i['order_status'] == 'Delivered')
        ordered = sum(1 for i in tab_items if i['order_status'] in ('Ordered', 'Delivered'))
        issues = sum(1 for i in tab_items if i['order_status'] == 'Issue')
        pct = round((delivered / total * 100) if total > 0 else 0)
        tab_data[tab] = {'categories': cats, 'total': total, 'delivered': delivered,
                         'ordered': ordered, 'issues': issues, 'pct': pct}

    total_all = len(items)
    ordered_all = sum(1 for i in items if i['order_status'] in ('Ordered', 'Delivered'))
    delivered_all = sum(1 for i in items if i['order_status'] == 'Delivered')
    issues_all = sum(1 for i in items if i['order_status'] == 'Issue')

    active_tab = request.args.get('tab', 'Office')
    return render_template('supplies.html',
        clinic=clinic, tabs=tabs, tab_data=tab_data,
        active_tab=active_tab,
        total_all=total_all, ordered_all=ordered_all,
        delivered_all=delivered_all, issues_all=issues_all,
        supply_statuses=SUPPLY_ORDER_STATUSES)

@app.route('/clinic/<int:clinic_id>/supplies/update', methods=['POST'])
@login_required
def update_supply_item(clinic_id):
    item_id = request.form.get('item_id')
    field = request.form.get('field')
    value = request.form.get('value', '')
    allowed_fields = ['order_status', 'expected_delivery_date', 'notes', 'amount_ordered', 'team_ordering']
    if field not in allowed_fields:
        return jsonify({'error': 'Invalid field'}), 400
    execute_db(f"UPDATE supply_items SET {field}=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND clinic_id=?",
               (value or None, item_id, clinic_id))
    return jsonify({'ok': True})

@app.route('/clinic/<int:clinic_id>/supplies/add', methods=['POST'])
@login_required
def add_supply_item(clinic_id):
    tab = request.form.get('tab', 'Post-Opening Orders')
    category = request.form.get('category', 'Additional')
    item_name = request.form.get('item_name', '').strip()
    if not item_name:
        flash('Item name required.', 'error')
        return redirect(url_for('clinic_supplies', clinic_id=clinic_id, tab=tab))
    team = request.form.get('team_ordering', '')
    qty = request.form.get('default_qty', '')
    execute_db(
        '''INSERT INTO supply_items (clinic_id, tab, category, item_name, default_qty, team_ordering, order_status)
           VALUES (?,?,?,?,?,?,?)''',
        (clinic_id, tab, category, item_name, qty, team, 'Not Ordered')
    )
    flash('Supply item added.', 'success')
    return redirect(url_for('clinic_supplies', clinic_id=clinic_id, tab=tab))

# ─── Construction Gantt ───────────────────────────────────────────────────────

@app.route('/clinic/<int:clinic_id>/construction')
@login_required
def clinic_construction(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    tasks = query_db("SELECT * FROM construction_tasks WHERE clinic_id=? ORDER BY order_index", [clinic_id])
    if not tasks:
        populate_construction_tasks(clinic_id)
        tasks = query_db("SELECT * FROM construction_tasks WHERE clinic_id=? ORDER BY order_index", [clinic_id])
    tasks_list = [dict(t) for t in tasks]
    return render_template('construction.html',
        clinic=clinic, tasks=tasks, tasks_json=tasks_list,
        construction_statuses=CONSTRUCTION_STATUSES)

@app.route('/clinic/<int:clinic_id>/construction/update', methods=['POST'])
@login_required
def update_construction_task(clinic_id):
    task_id = request.form.get('task_id')
    field = request.form.get('field')
    value = request.form.get('value', '')
    allowed = ['start_date', 'due_date', 'status', 'notes']
    if field not in allowed:
        return jsonify({'error': 'Invalid field'}), 400
    execute_db(f"UPDATE construction_tasks SET {field}=?, updated_at=CURRENT_TIMESTAMP WHERE id=? AND clinic_id=?",
               (value or None, task_id, clinic_id))
    return jsonify({'ok': True})

@app.route('/clinic/<int:clinic_id>/construction/add', methods=['POST'])
@login_required
def add_construction_task(clinic_id):
    name = request.form.get('name', '').strip()
    if not name:
        flash('Task name required.', 'error')
        return redirect(url_for('clinic_construction', clinic_id=clinic_id))
    max_order = query_db("SELECT MAX(order_index) as m FROM construction_tasks WHERE clinic_id=?", [clinic_id], one=True)
    next_order = (max_order['m'] or 0) + 1
    execute_db(
        "INSERT INTO construction_tasks (clinic_id, name, status, order_index) VALUES (?,?,?,?)",
        (clinic_id, name, 'Not Started', next_order)
    )
    flash('Construction task added.', 'success')
    return redirect(url_for('clinic_construction', clinic_id=clinic_id))

@app.route('/clinic/<int:clinic_id>/construction/delete/<int:task_id>', methods=['POST'])
@pm_required
def delete_construction_task(clinic_id, task_id):
    execute_db("DELETE FROM construction_tasks WHERE id=? AND clinic_id=?", (task_id, clinic_id))
    flash('Task deleted.', 'success')
    return redirect(url_for('clinic_construction', clinic_id=clinic_id))

@app.route('/clinic/<int:clinic_id>/construction/template')
@login_required
def construction_download_template(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('openpyxl not installed.', 'error')
        return redirect(url_for('clinic_construction', clinic_id=clinic_id))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Construction Timeline'

    headers = ['Task Name', 'Start Date', 'Due Date', 'Notes']
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, task_name in enumerate(CONSTRUCTION_TASKS_TEMPLATE, 2):
        ws.cell(row=row, column=1, value=task_name)

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    safe_name = clinic['name'].replace(' ', '_').replace('/', '_')
    filename = f'construction_timeline_{safe_name}.xlsx'
    return Response(
        out.read(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment;filename={filename}'}
    )


def _fuzzy_match_task(row_text, task_names):
    """Match row_text to one of task_names via keyword overlap."""
    row_lower = row_text.lower().strip()
    if not row_lower:
        return None
    best_match = None
    best_score = 0
    for task in task_names:
        task_lower = task.lower().strip()
        if task_lower == row_lower:
            return task
        keywords = [w for w in task_lower.split() if len(w) > 2]
        if not keywords:
            continue
        score = sum(1 for kw in keywords if kw in row_lower) / len(keywords)
        if score > best_score and score >= 0.4:
            best_score = score
            best_match = task
    return best_match


def _parse_date_str(val):
    """Try various date formats and return YYYY-MM-DD or None."""
    if not val:
        return None
    if hasattr(val, 'strftime'):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y', '%d/%m/%Y',
                '%B %d, %Y', '%B %d %Y', '%b %d, %Y', '%b %d %Y', '%m-%d-%Y']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return None


@app.route('/clinic/<int:clinic_id>/construction/upload', methods=['POST'])
@login_required
def upload_construction_timeline(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))

    f = request.files.get('timeline_file')
    if not f or not f.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('clinic_construction', clinic_id=clinic_id))

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ('xlsx', 'xls', 'pdf'):
        flash('Please upload a PDF or Excel (.xlsx/.xls) file.', 'error')
        return redirect(url_for('clinic_construction', clinic_id=clinic_id))

    matched = {}  # task_name -> {start_date, due_date}
    unmatched_rows = []

    if ext in ('xlsx', 'xls'):
        try:
            import openpyxl
            file_bytes = f.read()
            wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
            ws = wb.active

            # Detect header row and column positions
            header_row = 1
            task_col = start_col = end_col = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), 1):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        v = cell.value.lower().strip()
                        if any(kw in v for kw in ('task', 'name', 'description', 'activity', 'phase')):
                            task_col = cell.column
                            header_row = row_idx
                        elif any(kw in v for kw in ('start',)) and start_col is None:
                            start_col = cell.column
                        elif any(kw in v for kw in ('end', 'due', 'finish', 'complete', 'target')) and end_col is None:
                            end_col = cell.column
                if task_col:
                    break

            task_col = task_col or 1
            start_col = start_col or 2
            end_col = end_col or 3

            for row in ws.iter_rows(min_row=header_row + 1):
                def get_cell_val(col):
                    idx = col - 1
                    return row[idx].value if idx < len(row) else None

                row_text = str(get_cell_val(task_col) or '').strip()
                if not row_text:
                    continue
                start_date = _parse_date_str(get_cell_val(start_col))
                end_date = _parse_date_str(get_cell_val(end_col))
                match = _fuzzy_match_task(row_text, CONSTRUCTION_TASKS_TEMPLATE)
                if match:
                    matched[match] = {'start_date': start_date, 'due_date': end_date}
                else:
                    unmatched_rows.append(row_text)
        except Exception as e:
            flash(f'Error reading Excel file: {e}', 'error')
            return redirect(url_for('clinic_construction', clinic_id=clinic_id))

    elif ext == 'pdf':
        try:
            import fitz
            doc = fitz.open(stream=f.read(), filetype='pdf')
            text = ''.join(page.get_text() for page in doc)
            doc.close()
            date_pat = re.compile(
                r'\b(\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2}|\w+ \d{1,2},?\s*\d{4})\b'
            )
            lines = [l.strip() for l in text.split('\n')]
            for i, line in enumerate(lines):
                if not line:
                    continue
                match = _fuzzy_match_task(line, CONSTRUCTION_TASKS_TEMPLATE)
                if match:
                    context = ' '.join(lines[max(0, i - 1):i + 3])
                    dates = date_pat.findall(context)
                    matched[match] = {
                        'start_date': _parse_date_str(dates[0]) if len(dates) > 0 else None,
                        'due_date': _parse_date_str(dates[1]) if len(dates) > 1 else None,
                    }
        except Exception as e:
            flash(f'Error reading PDF file: {e}', 'error')
            return redirect(url_for('clinic_construction', clinic_id=clinic_id))

    # Update database
    for task_name, dates in matched.items():
        task_row = query_db(
            "SELECT id FROM construction_tasks WHERE clinic_id=? AND name=?",
            [clinic_id, task_name], one=True
        )
        if task_row:
            execute_db(
                '''UPDATE construction_tasks SET start_date=?, due_date=?,
                   updated_at=CURRENT_TIMESTAMP WHERE id=?''',
                (dates['start_date'], dates['due_date'], task_row['id'])
            )

    matched_count = len(matched)
    total_tasks = len(CONSTRUCTION_TASKS_TEMPLATE)
    unmatched_names = [t for t in CONSTRUCTION_TASKS_TEMPLATE if t not in matched]
    msg = f'Matched {matched_count}/{total_tasks} tasks.'
    if unmatched_names:
        display = unmatched_names[:5]
        msg += f' Unmatched: {", ".join(display)}'
        if len(unmatched_names) > 5:
            msg += f' and {len(unmatched_names) - 5} more.'
    flash(msg, 'success' if matched_count > 0 else 'warning')
    return redirect(url_for('clinic_construction', clinic_id=clinic_id))

# ─── Status Report & Due This Week ────────────────────────────────────────────

@app.route('/clinic/<int:clinic_id>/status-report')
@login_required
def status_report(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    tasks = query_db(
        "SELECT * FROM tasks WHERE clinic_id=? AND status != 'Complete' ORDER BY department, time_phase, due_date",
        [clinic_id]
    )
    today = datetime.now().date()
    week_end = today + timedelta(days=7)

    dept_tasks = {}
    for dept in DEPARTMENTS:
        dt = [t for t in tasks if t['department'] == dept]
        if dt:
            dept_tasks[dept] = dt

    return render_template('status_report.html',
        clinic=clinic, dept_tasks=dept_tasks,
        today=str(today), week_end=str(week_end))

@app.route('/due-this-week')
@login_required
def due_this_week():
    today = datetime.now().date()
    week_end = today + timedelta(days=7)
    tasks = query_db(
        '''SELECT t.*, c.name as clinic_name FROM tasks t
           JOIN clinics c ON t.clinic_id = c.id
           WHERE c.is_template=0 AND t.due_date >= ? AND t.due_date <= ?
           AND t.status != 'Complete'
           ORDER BY t.due_date, c.name''',
        [str(today), str(week_end)]
    )
    overdue = query_db(
        '''SELECT t.*, c.name as clinic_name FROM tasks t
           JOIN clinics c ON t.clinic_id = c.id
           WHERE c.is_template=0 AND t.due_date < ?
           AND t.status != 'Complete'
           ORDER BY t.due_date, c.name''',
        [str(today)]
    )
    return render_template('due_this_week.html',
        tasks=tasks, overdue=overdue,
        today=str(today), week_end=str(week_end))

# ─── Pipeline Tracker ─────────────────────────────────────────────────────────

@app.route('/pipeline')
@login_required
def pipeline():
    status_filter = request.args.get('status', '')
    pm_filter = request.args.get('pm', '')
    state_filter = request.args.get('state', '')

    q = "SELECT * FROM clinics WHERE is_template=0"
    args = []
    if status_filter:
        q += " AND site_status=?"; args.append(status_filter)
    if pm_filter:
        q += " AND sg_project_manager=?"; args.append(pm_filter)
    if state_filter:
        q += " AND state=?"; args.append(state_filter)
    q += " ORDER BY opening_date ASC NULLS LAST, name ASC"
    clinics = query_db(q, args)

    quarters = {}
    for c in clinics:
        q_label = get_quarter(c['opening_date']) or 'No Date Set'
        if q_label not in quarters:
            quarters[q_label] = []
        quarters[q_label].append(dict(c))

    def quarter_sort_key(k):
        if k == 'No Date Set':
            return '9999'
        parts = k.split(' ')
        return f"{parts[1]}_{parts[0]}" if len(parts) == 2 else k
    sorted_quarters = dict(sorted(quarters.items(), key=lambda x: quarter_sort_key(x[0])))

    all_clinics_raw = query_db("SELECT DISTINCT site_status, sg_project_manager, state FROM clinics WHERE is_template=0")
    statuses_avail = sorted(set(c['site_status'] for c in all_clinics_raw if c['site_status']))
    pms_avail = sorted(set(c['sg_project_manager'] for c in all_clinics_raw if c['sg_project_manager']))
    states_avail = sorted(set(c['state'] for c in all_clinics_raw if c['state']))

    return render_template('pipeline.html',
        quarters=sorted_quarters, clinics=clinics,
        statuses_avail=statuses_avail, pms_avail=pms_avail, states_avail=states_avail,
        status_filter=status_filter, pm_filter=pm_filter, state_filter=state_filter)

@app.route('/pipeline/export-csv')
@login_required
def pipeline_export_csv():
    clinics = query_db("SELECT * FROM clinics WHERE is_template=0 ORDER BY opening_date ASC NULLS LAST, name ASC")
    si = StringIO()
    writer = csv.writer(si)
    writer.writerow(['Clinic Name', 'State', 'Entity', 'PM', 'Site Status', 'Set Up Week',
                     'Doctor Status', 'Opening Date', 'Ops Onsite Member', 'Targeting Month',
                     'Total Buildout Cost', 'Cost to VIP', 'Notes'])
    for c in clinics:
        writer.writerow([
            c['name'], c['state'] or '', c['entity'] or '', c['sg_project_manager'] or '',
            c['site_status'] or '', c['setup_week'] or '', c['doctor_status'] or '',
            c['opening_date'] or '', c['ops_onsite_member'] or '',
            c['targeting_opening_month'] or '', c['total_buildout_cost'] or '',
            c['cost_to_vip'] or '', c['clinic_notes'] or ''
        ])
    output = si.getvalue()
    return Response(output, mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=pipeline_tracker.csv'})

# ─── Notifications & misc ─────────────────────────────────────────────────────

@app.route('/notifications')
@login_required
def notifications():
    notifs = query_db(
        "SELECT * FROM notifications WHERE user_id=? ORDER BY created_at DESC LIMIT 50",
        [session['user_id']])
    execute_db("UPDATE notifications SET is_read=1 WHERE user_id=?", [session['user_id']])
    return render_template('notifications.html', notifications=notifs)

@app.route('/api/notifications/count')
@login_required
def notif_count():
    row = query_db("SELECT COUNT(*) as cnt FROM notifications WHERE user_id=? AND is_read=0",
                   [session['user_id']], one=True)
    return jsonify({'count': row['cnt'] if row else 0})

@app.route('/clinic/<int:clinic_id>/quickcheck')
@login_required
def quick_check(clinic_id):
    dept = request.args.get('dept', DEPARTMENTS[0])
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    recent_tasks = query_db(
        '''SELECT t.*, n.content as last_note, n.created_at as note_time, u.full_name as note_author
           FROM tasks t
           LEFT JOIN notes n ON n.id = (SELECT id FROM notes WHERE task_id=t.id ORDER BY created_at DESC LIMIT 1)
           LEFT JOIN users u ON n.author_id=u.id
           WHERE t.clinic_id=? AND t.department=?
           ORDER BY t.updated_at DESC LIMIT 10''',
        [clinic_id, dept])
    return render_template('quick_check.html', clinic=clinic, dept=dept,
                           tasks=recent_tasks, departments=DEPARTMENTS)

@app.route('/admin/users')
@admin_required
def admin_users():
    users = query_db("SELECT * FROM users ORDER BY full_name")
    return render_template('admin_users.html', users=users, departments=DEPARTMENTS)

@app.route('/admin/users/new', methods=['POST'])
@admin_required
def admin_new_user():
    username = request.form.get('username', '').strip()
    full_name = request.form.get('full_name', '').strip()
    password = request.form.get('password', '')
    role = request.form.get('role', 'team_member')
    dept = request.form.get('department', '') or None
    if not username or not full_name or not password:
        flash('All fields required.', 'error')
        return redirect(url_for('admin_users'))
    existing = query_db("SELECT id FROM users WHERE username=?", [username], one=True)
    if existing:
        flash('Username already exists.', 'error')
        return redirect(url_for('admin_users'))
    execute_db(
        "INSERT INTO users (username, password_hash, full_name, role, department) VALUES (?,?,?,?,?)",
        (username, generate_password_hash(password), full_name, role, dept)
    )
    flash(f'User {full_name} created.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@admin_required
def admin_edit_user(user_id):
    user = query_db("SELECT * FROM users WHERE id=?", [user_id], one=True)
    if not user:
        flash('User not found.', 'error')
        return redirect(url_for('admin_users'))
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip() or None
        role = request.form.get('role', 'team_member')
        department = request.form.get('department', '') or None
        new_password = request.form.get('new_password', '').strip()
        if not full_name or not username:
            flash('Full name and username are required.', 'error')
            return render_template('admin_edit_user.html', user=user, departments=DEPARTMENTS)
        # Check username uniqueness (excluding this user)
        existing = query_db("SELECT id FROM users WHERE username=? AND id != ?", [username, user_id], one=True)
        if existing:
            flash('Username already taken.', 'error')
            return render_template('admin_edit_user.html', user=user, departments=DEPARTMENTS)
        if new_password:
            execute_db(
                "UPDATE users SET full_name=?, username=?, role=?, department=?, password_hash=? WHERE id=?",
                (full_name, username, role, department, generate_password_hash(new_password), user_id)
            )
        else:
            execute_db(
                "UPDATE users SET full_name=?, username=?, role=?, department=? WHERE id=?",
                (full_name, username, role, department, user_id)
            )
        # Update session if editing self
        if user_id == session['user_id']:
            session['full_name'] = full_name
            session['role'] = role
        flash(f'User {full_name} updated successfully.', 'success')
        return redirect(url_for('admin_users'))
    return render_template('admin_edit_user.html', user=user, departments=DEPARTMENTS)

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    if user_id == session['user_id']:
        flash('Cannot delete yourself.', 'error')
        return redirect(url_for('admin_users'))
    execute_db("DELETE FROM users WHERE id=?", [user_id])
    flash('User deleted.', 'success')
    return redirect(url_for('admin_users'))

@app.route('/template')
@admin_required
def template_tasks():
    tmpl = query_db("SELECT * FROM clinics WHERE is_template=1", one=True)
    if not tmpl:
        flash('Template not found.', 'error')
        return redirect(url_for('index'))
    tasks = query_db("SELECT * FROM tasks WHERE clinic_id=? ORDER BY department, template_offset_days, order_index", [tmpl['id']])
    users_list = query_db("SELECT id, full_name FROM users ORDER BY full_name")
    return render_template('template.html', clinic=tmpl, tasks=tasks, users=users_list,
                           departments=DEPARTMENTS, time_phases=TIME_PHASES, statuses=STATUSES)

@app.route('/template/task/new', methods=['POST'])
@admin_required
def template_new_task():
    tmpl = query_db("SELECT * FROM clinics WHERE is_template=1", one=True)
    name = request.form.get('name', '').strip()
    dept = request.form.get('department', '')
    phase = request.form.get('time_phase', '')
    offset = int(request.form.get('template_offset_days', 0))
    if name and dept:
        execute_db(
            '''INSERT INTO tasks (clinic_id, name, department, time_phase, is_template, template_offset_days)
               VALUES (?,?,?,?,?,?)''',
            (tmpl['id'], name, dept, phase, 1, offset)
        )
        flash('Template task added.', 'success')
    return redirect(url_for('template_tasks'))

@app.route('/template/task/<int:task_id>/delete', methods=['POST'])
@admin_required
def template_delete_task(task_id):
    execute_db("DELETE FROM tasks WHERE id=?", [task_id])
    flash('Template task deleted.', 'success')
    return redirect(url_for('template_tasks'))

@app.route('/clinic/<int:clinic_id>/delete', methods=['POST'])
@admin_required
def delete_clinic(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if clinic:
        execute_db("DELETE FROM notes WHERE task_id IN (SELECT id FROM tasks WHERE clinic_id=?)", [clinic_id])
        execute_db("DELETE FROM attachments WHERE task_id IN (SELECT id FROM tasks WHERE clinic_id=?)", [clinic_id])
        execute_db("DELETE FROM tasks WHERE clinic_id=?", [clinic_id])
        execute_db("DELETE FROM activity_log WHERE clinic_id=?", [clinic_id])
        execute_db("DELETE FROM lease_summaries WHERE clinic_id=?", [clinic_id])
        execute_db("DELETE FROM supply_items WHERE clinic_id=?", [clinic_id])
        execute_db("DELETE FROM construction_tasks WHERE clinic_id=?", [clinic_id])
        execute_db("DELETE FROM clinics WHERE id=?", [clinic_id])
        flash(f'Clinic "{clinic["name"]}" deleted.', 'success')
    return redirect(url_for('index'))

# ─── Expense Tracker ──────────────────────────────────────────────────────────

@app.route('/clinic/<int:clinic_id>/expenses')
@login_required
def clinic_expenses(clinic_id):
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    expenses = query_db("SELECT * FROM expenses WHERE clinic_id=? ORDER BY category, id", [clinic_id])
    if not expenses:
        populate_expense_templates(clinic_id, session['user_id'])
        expenses = query_db("SELECT * FROM expenses WHERE clinic_id=? ORDER BY category, id", [clinic_id])

    # Group by category
    categories = {}
    for exp in expenses:
        cat = exp['category']
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(dict(exp))

    total_expenses = sum(e['amount'] or 0 for e in expenses)
    return render_template('expenses.html',
        clinic=clinic, categories=categories, expenses=expenses,
        total_expenses=total_expenses)

@app.route('/clinic/<int:clinic_id>/expenses/update', methods=['POST'])
@login_required
def update_expense(clinic_id):
    expense_id = request.form.get('expense_id')
    field = request.form.get('field')
    value = request.form.get('value', '')
    allowed = ['expense_date', 'description', 'vendor', 'amount', 'category']
    if field not in allowed:
        return jsonify({'error': 'Invalid field'}), 400
    if field == 'amount':
        try:
            value = float(value) if value else 0
        except ValueError:
            value = 0
    execute_db(f"UPDATE expenses SET {field}=? WHERE id=? AND clinic_id=?",
               (value or None if field != 'amount' else value, expense_id, clinic_id))
    return jsonify({'ok': True})

@app.route('/clinic/<int:clinic_id>/expenses/add', methods=['POST'])
@login_required
def add_expense(clinic_id):
    category = request.form.get('category', '').strip()
    description = request.form.get('description', '').strip()
    vendor = request.form.get('vendor', '').strip()
    if not category:
        return jsonify({'error': 'Category required'}), 400
    new_id = execute_db(
        '''INSERT INTO expenses (clinic_id, category, description, vendor, amount, created_by)
           VALUES (?,?,?,?,0,?)''',
        (clinic_id, category, description, vendor, session['user_id'])
    )
    return jsonify({'ok': True, 'id': new_id})

@app.route('/clinic/<int:clinic_id>/expenses/delete/<int:expense_id>', methods=['POST'])
@login_required
def delete_expense(clinic_id, expense_id):
    execute_db("DELETE FROM expenses WHERE id=? AND clinic_id=?", (expense_id, clinic_id))
    return jsonify({'ok': True})

@app.route('/clinic/<int:clinic_id>/expenses/summary', methods=['POST'])
@login_required
def update_expense_summary(clinic_id):
    buildout_cost = request.form.get('buildout_cost', '') or None
    ti_allowance = request.form.get('ti_allowance', '') or None
    ti_paid_by_vip = request.form.get('ti_paid_by_vip', '') or None
    try:
        buildout_cost = float(buildout_cost) if buildout_cost else None
    except ValueError:
        buildout_cost = None
    try:
        ti_allowance = float(ti_allowance) if ti_allowance else None
    except ValueError:
        ti_allowance = None
    try:
        ti_paid_by_vip = float(ti_paid_by_vip) if ti_paid_by_vip else None
    except ValueError:
        ti_paid_by_vip = None
    execute_db("UPDATE clinics SET buildout_cost=?, ti_allowance=?, ti_paid_by_vip=? WHERE id=?",
               (buildout_cost, ti_allowance, ti_paid_by_vip, clinic_id))
    flash('Summary updated.', 'success')
    return redirect(url_for('clinic_expenses', clinic_id=clinic_id))

# ─── Weekly Leadership Report ─────────────────────────────────────────────────

def get_week_ending(date=None):
    """Return the Friday of the current (or given) week as YYYY-MM-DD."""
    if date is None:
        date = datetime.now().date()
    elif isinstance(date, str):
        date = datetime.strptime(date, '%Y-%m-%d').date()
    # weekday(): Monday=0, Friday=4
    days_until_friday = (4 - date.weekday()) % 7
    friday = date + timedelta(days=days_until_friday)
    return str(friday)

@app.route('/weekly-report')
@login_required
def weekly_report_all():
    week_ending = request.args.get('week', get_week_ending())
    # Generate a list of recent Fridays for the dropdown
    today = datetime.now().date()
    weeks = []
    for i in range(-4, 13):
        d = today + timedelta(weeks=i)
        we = get_week_ending(d)
        if we not in weeks:
            weeks.append(we)
    weeks = sorted(set(weeks))

    clinics = query_db("SELECT * FROM clinics WHERE is_template=0 ORDER BY opening_date ASC NULLS LAST, name ASC")
    clinic_reports = []
    for clinic in clinics:
        report = query_db("SELECT * FROM weekly_reports WHERE clinic_id=? AND week_ending=?",
                          [clinic['id'], week_ending], one=True)
        # Compute stats
        total = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=?", [clinic['id']], one=True)['cnt']
        done = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND status='Complete'", [clinic['id']], one=True)['cnt']
        pct = round((done / total * 100) if total > 0 else 0)

        # Supply %
        s_total = query_db("SELECT COUNT(*) as cnt FROM supply_items WHERE clinic_id=?", [clinic['id']], one=True)['cnt']
        s_delivered = query_db("SELECT COUNT(*) as cnt FROM supply_items WHERE clinic_id=? AND order_status='Delivered'", [clinic['id']], one=True)['cnt']
        supply_pct = round((s_delivered / s_total * 100) if s_total > 0 else 0)

        # Construction current phase
        current_construction = query_db(
            "SELECT name FROM construction_tasks WHERE clinic_id=? AND status='In Progress' ORDER BY order_index LIMIT 1",
            [clinic['id']], one=True)
        if not current_construction:
            current_construction = query_db(
                "SELECT name FROM construction_tasks WHERE clinic_id=? AND status='Not Started' ORDER BY order_index LIMIT 1",
                [clinic['id']], one=True)
        construction_phase = current_construction['name'] if current_construction else 'Complete'

        clinic_reports.append({
            'clinic': dict(clinic),
            'report': dict(report) if report else None,
            'pct': pct,
            'supply_pct': supply_pct,
            'construction_phase': construction_phase,
        })
    return render_template('weekly_report_all.html',
        clinic_reports=clinic_reports,
        week_ending=week_ending,
        weeks=weeks)

@app.route('/clinic/<int:clinic_id>/weekly-report')
@login_required
def weekly_report_view(clinic_id):
    week_ending = request.args.get('week', get_week_ending())
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))
    report = query_db("SELECT * FROM weekly_reports WHERE clinic_id=? AND week_ending=?",
                      [clinic_id, week_ending], one=True)
    total = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=?", [clinic_id], one=True)['cnt']
    done = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND status='Complete'", [clinic_id], one=True)['cnt']
    pct = round((done / total * 100) if total > 0 else 0)
    s_total = query_db("SELECT COUNT(*) as cnt FROM supply_items WHERE clinic_id=?", [clinic_id], one=True)['cnt']
    s_delivered = query_db("SELECT COUNT(*) as cnt FROM supply_items WHERE clinic_id=? AND order_status='Delivered'", [clinic_id], one=True)['cnt']
    supply_pct = round((s_delivered / s_total * 100) if s_total > 0 else 0)
    current_construction = query_db(
        "SELECT name FROM construction_tasks WHERE clinic_id=? AND status='In Progress' ORDER BY order_index LIMIT 1",
        [clinic_id], one=True)
    if not current_construction:
        current_construction = query_db(
            "SELECT name FROM construction_tasks WHERE clinic_id=? AND status='Not Started' ORDER BY order_index LIMIT 1",
            [clinic_id], one=True)
    construction_phase = current_construction['name'] if current_construction else 'Complete'
    return render_template('weekly_report_view.html',
        clinic=clinic, report=report, pct=pct,
        supply_pct=supply_pct, construction_phase=construction_phase,
        week_ending=week_ending)

@app.route('/clinic/<int:clinic_id>/weekly-report/edit', methods=['GET', 'POST'])
@login_required
def weekly_report_edit(clinic_id):
    week_ending = request.args.get('week', get_week_ending())
    clinic = query_db("SELECT * FROM clinics WHERE id=?", [clinic_id], one=True)
    if not clinic:
        flash('Clinic not found.', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        week_ending = request.form.get('week_ending', week_ending)
        status = request.form.get('status', 'On Track')
        highlights = request.form.get('highlights', '') or None
        next_week = request.form.get('next_week', '') or None
        timeline_changes = request.form.get('timeline_changes', '') or None
        approvals_needed = request.form.get('approvals_needed', '') or None
        help_needed = request.form.get('help_needed', '') or None

        existing = query_db("SELECT id FROM weekly_reports WHERE clinic_id=? AND week_ending=?",
                            [clinic_id, week_ending], one=True)
        if existing:
            execute_db(
                '''UPDATE weekly_reports SET status=?, highlights=?, next_week=?,
                   timeline_changes=?, approvals_needed=?, help_needed=? WHERE id=?''',
                (status, highlights, next_week, timeline_changes, approvals_needed, help_needed, existing['id'])
            )
        else:
            execute_db(
                '''INSERT INTO weekly_reports (clinic_id, week_ending, status, highlights, next_week,
                   timeline_changes, approvals_needed, help_needed, created_by)
                   VALUES (?,?,?,?,?,?,?,?,?)''',
                (clinic_id, week_ending, status, highlights, next_week,
                 timeline_changes, approvals_needed, help_needed, session['user_id'])
            )
        flash('Weekly report saved.', 'success')
        return redirect(url_for('weekly_report_view', clinic_id=clinic_id, week=week_ending))

    report = query_db("SELECT * FROM weekly_reports WHERE clinic_id=? AND week_ending=?",
                      [clinic_id, week_ending], one=True)
    return render_template('weekly_report_edit.html',
        clinic=clinic, report=report, week_ending=week_ending)

# ─── Reports ─────────────────────────────────────────────────────────────────

@app.route('/reports/leases')
@login_required
def reports_leases():
    clinics = query_db(
        "SELECT * FROM clinics WHERE is_template=0 AND status != 'Archived' ORDER BY name ASC"
    )
    rows = []
    for c in clinics:
        lease = query_db("SELECT * FROM lease_summaries WHERE clinic_id=?", [c['id']], one=True)
        # Compute net cost to VIP from lease data
        net_cost_vip = None
        if lease:
            def parse_dollar(s):
                if not s:
                    return None
                try:
                    return float(str(s).replace('$', '').replace(',', '').strip())
                except Exception:
                    return None
            buildout = parse_dollar(lease['buildout_cost_estimate'] if lease else None)
            ti = parse_dollar(lease['ti_allowance'] if lease else None)
            if buildout is not None and ti is not None:
                net_cost_vip = buildout - ti
            elif buildout is not None:
                net_cost_vip = buildout
        rows.append({'clinic': dict(c), 'lease': dict(lease) if lease else None, 'net_cost_vip': net_cost_vip})
    return render_template('reports_leases.html', rows=rows)


@app.route('/reports/builder/save', methods=['POST'])
@login_required
def save_report_template():
    name = request.form.get('name', '').strip()
    fields_json = request.form.get('fields', '[]')
    try:
        fields = json.loads(fields_json)
    except Exception:
        fields = request.form.getlist('fields')
    if not name or not fields:
        return jsonify({'error': 'Name and fields required'}), 400
    new_id = execute_db(
        "INSERT INTO report_templates (name, fields, created_by) VALUES (?,?,?)",
        (name, json.dumps(fields), session['user_id'])
    )
    return jsonify({'ok': True, 'id': new_id, 'name': name, 'fields': fields})


@app.route('/reports/builder/templates')
@login_required
def list_report_templates():
    templates = query_db("SELECT * FROM report_templates ORDER BY created_at DESC")
    result = [{'id': t['id'], 'name': t['name'], 'fields': json.loads(t['fields'] or '[]'),
               'created_at': t['created_at']} for t in templates]
    return jsonify(result)


@app.route('/reports/builder/template/<int:template_id>', methods=['DELETE'])
@login_required
def delete_report_template(template_id):
    execute_db("DELETE FROM report_templates WHERE id=?", [template_id])
    return jsonify({'ok': True})


@app.route('/reports/builder')
@login_required
def reports_builder():
    # Get selected fields from URL params
    selected_fields = request.args.getlist('fields')

    # All field definitions: key -> (label, source)
    FIELD_GROUPS = [
        ('Basic Info', [
            ('clinic_name', 'Clinic Name'),
            ('state', 'State'),
            ('entity', 'Entity'),
            ('site_status', 'Site Status'),
            ('doctor_status', 'Doctor Status'),
            ('targeting_opening_month', 'Targeting Opening Month'),
            ('opening_date', 'Opening Date'),
            ('setup_week', 'Setup Week'),
            ('procedures_done_where', 'Procedures Done Where'),
            ('paired_clinic', 'Paired Clinic'),
        ]),
        ('Team', [
            ('sg_project_manager', 'SG Project Manager'),
            ('sg_onsite_member', 'SG Onsite Member'),
            ('ops_onsite_member', 'Operations Onsite Member'),
        ]),
        ('Lease', [
            ('lease_signed_date', 'Lease Signed Date'),
            ('building_address', 'Building Address'),
            ('suite_unit', 'Suite/Unit'),
            ('square_footage', 'Square Footage'),
            ('landlord_name', 'Landlord Name'),
            ('lease_start_date', 'Lease Start'),
            ('lease_end_date', 'Lease End'),
            ('base_rent', 'Base Rent'),
            ('cam_rent', 'CAMs'),
            ('total_monthly_rent', 'Total Monthly Rent'),
            ('annual_rent_increase_pct', 'Annual Increase %'),
            ('security_deposit', 'Security Deposit'),
            ('ti_allowance', 'TI Allowance'),
            ('buildout_cost_estimate', 'Buildout Cost Estimate'),
            ('lease_net_cost_vip', 'Net Cost to VIP'),
            ('attorney_name', 'Attorney Name'),
        ]),
        ('Progress', [
            ('overall_pct', 'Overall % Complete'),
            ('tasks_complete', 'Tasks Complete'),
            ('tasks_total', 'Tasks Total'),
            ('tasks_overdue', 'Tasks Overdue'),
            ('construction_phase', 'Construction Phase'),
            ('supplies_pct', 'Supplies % Delivered'),
        ]),
        ('Financials', [
            ('total_expenses', 'Total Expenses'),
            ('buildout_cost', 'Buildout Cost (from clinics)'),
            ('ti_allowance_clinic', 'TI Allowance (from clinics)'),
            ('net_cost_vip_clinic', 'Net Cost to VIP (clinics)'),
        ]),
        ('Notes', [
            ('data_analysis_summary', 'Data Analysis Summary'),
            ('clinic_notes', 'Clinic Notes'),
        ]),
    ]

    report_rows = []
    if selected_fields:
        clinics = query_db(
            "SELECT * FROM clinics WHERE is_template=0 AND status != 'Archived' ORDER BY name ASC"
        )
        today_str = str(datetime.now().date())

        for c in clinics:
            lease = query_db("SELECT * FROM lease_summaries WHERE clinic_id=?", [c['id']], one=True)

            def parse_dollar(s):
                if not s:
                    return None
                try:
                    return float(str(s).replace('$', '').replace(',', '').strip())
                except Exception:
                    return None

            # Compute progress stats
            total_tasks = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=?", [c['id']], one=True)['cnt']
            done_tasks = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND status='Complete'", [c['id']], one=True)['cnt']
            overdue_tasks = query_db(
                "SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND due_date < ? AND status NOT IN ('Complete')",
                [c['id'], today_str], one=True)['cnt']
            pct = round((done_tasks / total_tasks * 100) if total_tasks > 0 else 0)

            s_total = query_db("SELECT COUNT(*) as cnt FROM supply_items WHERE clinic_id=?", [c['id']], one=True)['cnt']
            s_delivered = query_db("SELECT COUNT(*) as cnt FROM supply_items WHERE clinic_id=? AND order_status='Delivered'", [c['id']], one=True)['cnt']
            supply_pct = round((s_delivered / s_total * 100) if s_total > 0 else 0)

            current_construction = query_db(
                "SELECT name FROM construction_tasks WHERE clinic_id=? AND status='In Progress' ORDER BY order_index LIMIT 1",
                [c['id']], one=True)
            if not current_construction:
                current_construction = query_db(
                    "SELECT name FROM construction_tasks WHERE clinic_id=? AND status='Not Started' ORDER BY order_index LIMIT 1",
                    [c['id']], one=True)
            construction_phase = current_construction['name'] if current_construction else 'Complete'

            total_expenses = query_db(
                "SELECT SUM(amount) as s FROM expenses WHERE clinic_id=?", [c['id']], one=True)['s'] or 0

            # Net cost to VIP from lease
            lease_net_cost_vip = None
            if lease:
                buildout = parse_dollar(lease['buildout_cost_estimate'])
                ti = parse_dollar(lease['ti_allowance'])
                if buildout is not None and ti is not None:
                    lease_net_cost_vip = buildout - ti
                elif buildout is not None:
                    lease_net_cost_vip = buildout

            # Net cost to VIP from clinics table
            net_cost_vip_clinic = None
            bc = parse_dollar(c['buildout_cost']) if 'buildout_cost' in c.keys() else None
            ti_c = parse_dollar(c['ti_allowance']) if 'ti_allowance' in c.keys() else None
            if bc is not None and ti_c is not None:
                net_cost_vip_clinic = bc - ti_c

            def fmt_money(n):
                if n is None:
                    return ''
                return f'${n:,.0f}'

            # Build row dict
            all_data = {
                'clinic_name': c['name'],
                'state': c['state'] or '',
                'entity': c['entity'] or '',
                'site_status': c['site_status'] or '',
                'doctor_status': c['doctor_status'] or '',
                'targeting_opening_month': c['targeting_opening_month'] or '',
                'opening_date': c['opening_date'] or '',
                'setup_week': c['setup_week'] or '',
                'procedures_done_where': c['procedures_done_where'] or '',
                'paired_clinic': c['paired_clinic'] or '',
                'sg_project_manager': c['sg_project_manager'] or '',
                'sg_onsite_member': c['sg_onsite_member'] or '',
                'ops_onsite_member': c['ops_onsite_member'] or '',
                'lease_signed_date': c['lease_signed_date'] or '',
                'building_address': (lease['building_address'] if lease else '') or '',
                'suite_unit': (lease['suite_unit'] if lease else '') or '',
                'square_footage': (lease['square_footage'] if lease else '') or '',
                'landlord_name': (lease['landlord_name'] if lease else '') or '',
                'lease_start_date': (lease['lease_start_date'] if lease else '') or '',
                'lease_end_date': (lease['lease_end_date'] if lease else '') or '',
                'base_rent': (lease['base_rent'] if lease else '') or '',
                'cam_rent': (lease['cam_rent'] if lease else '') or '',
                'total_monthly_rent': (lease['total_monthly_rent'] if lease else '') or '',
                'annual_rent_increase_pct': (lease['annual_rent_increase_pct'] if lease else '') or '',
                'security_deposit': (lease['security_deposit'] if lease else '') or '',
                'ti_allowance': (lease['ti_allowance'] if lease else '') or '',
                'buildout_cost_estimate': (lease['buildout_cost_estimate'] if lease else '') or '',
                'lease_net_cost_vip': fmt_money(lease_net_cost_vip),
                'attorney_name': (lease['attorney_name'] if lease else '') or '',
                'overall_pct': f'{pct}%',
                'tasks_complete': str(done_tasks),
                'tasks_total': str(total_tasks),
                'tasks_overdue': str(overdue_tasks),
                'construction_phase': construction_phase,
                'supplies_pct': f'{supply_pct}%',
                'total_expenses': fmt_money(total_expenses),
                'buildout_cost': fmt_money(bc),
                'ti_allowance_clinic': fmt_money(ti_c),
                'net_cost_vip_clinic': fmt_money(net_cost_vip_clinic),
                'data_analysis_summary': c['data_analysis_summary'] or '',
                'clinic_notes': c['clinic_notes'] or '',
            }

            row = {f: all_data.get(f, '') for f in selected_fields}
            report_rows.append(row)

    saved_templates = query_db("SELECT * FROM report_templates ORDER BY created_at DESC")
    saved_templates_list = [{'id': t['id'], 'name': t['name'],
                              'fields': json.loads(t['fields'] or '[]')} for t in saved_templates]
    return render_template('reports_builder.html',
        field_groups=FIELD_GROUPS,
        selected_fields=selected_fields,
        report_rows=report_rows,
        saved_templates=saved_templates_list)


@app.route('/reports/builder/csv')
@login_required
def reports_builder_csv():
    selected_fields = request.args.getlist('fields')
    if not selected_fields:
        return redirect(url_for('reports_builder'))

    FIELD_LABELS = {
        'clinic_name': 'Clinic Name', 'state': 'State', 'entity': 'Entity',
        'site_status': 'Site Status', 'doctor_status': 'Doctor Status',
        'targeting_opening_month': 'Targeting Opening Month', 'opening_date': 'Opening Date',
        'setup_week': 'Setup Week', 'procedures_done_where': 'Procedures Done Where',
        'paired_clinic': 'Paired Clinic', 'sg_project_manager': 'SG Project Manager',
        'sg_onsite_member': 'SG Onsite Member', 'ops_onsite_member': 'Operations Onsite Member',
        'lease_signed_date': 'Lease Signed Date', 'building_address': 'Building Address',
        'suite_unit': 'Suite/Unit', 'square_footage': 'Square Footage',
        'landlord_name': 'Landlord Name', 'lease_start_date': 'Lease Start',
        'lease_end_date': 'Lease End', 'base_rent': 'Base Rent', 'cam_rent': 'CAMs',
        'total_monthly_rent': 'Total Monthly Rent', 'annual_rent_increase_pct': 'Annual Increase %',
        'security_deposit': 'Security Deposit', 'ti_allowance': 'TI Allowance',
        'buildout_cost_estimate': 'Buildout Cost Estimate', 'lease_net_cost_vip': 'Net Cost to VIP',
        'attorney_name': 'Attorney Name', 'overall_pct': 'Overall % Complete',
        'tasks_complete': 'Tasks Complete', 'tasks_total': 'Tasks Total',
        'tasks_overdue': 'Tasks Overdue', 'construction_phase': 'Construction Phase',
        'supplies_pct': 'Supplies % Delivered', 'total_expenses': 'Total Expenses',
        'buildout_cost': 'Buildout Cost (clinics)', 'ti_allowance_clinic': 'TI Allowance (clinics)',
        'net_cost_vip_clinic': 'Net Cost to VIP (clinics)',
        'data_analysis_summary': 'Data Analysis Summary', 'clinic_notes': 'Clinic Notes',
    }

    clinics = query_db(
        "SELECT * FROM clinics WHERE is_template=0 AND status != 'Archived' ORDER BY name ASC"
    )
    today_str = str(datetime.now().date())

    si = StringIO()
    writer = csv.writer(si)
    writer.writerow([FIELD_LABELS.get(f, f) for f in selected_fields])

    for c in clinics:
        lease = query_db("SELECT * FROM lease_summaries WHERE clinic_id=?", [c['id']], one=True)

        def parse_dollar(s):
            if not s:
                return None
            try:
                return float(str(s).replace('$', '').replace(',', '').strip())
            except Exception:
                return None

        total_tasks = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=?", [c['id']], one=True)['cnt']
        done_tasks = query_db("SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND status='Complete'", [c['id']], one=True)['cnt']
        overdue_tasks = query_db(
            "SELECT COUNT(*) as cnt FROM tasks WHERE clinic_id=? AND due_date < ? AND status NOT IN ('Complete')",
            [c['id'], today_str], one=True)['cnt']
        pct = round((done_tasks / total_tasks * 100) if total_tasks > 0 else 0)

        s_total = query_db("SELECT COUNT(*) as cnt FROM supply_items WHERE clinic_id=?", [c['id']], one=True)['cnt']
        s_delivered = query_db("SELECT COUNT(*) as cnt FROM supply_items WHERE clinic_id=? AND order_status='Delivered'", [c['id']], one=True)['cnt']
        supply_pct = round((s_delivered / s_total * 100) if s_total > 0 else 0)

        current_construction = query_db(
            "SELECT name FROM construction_tasks WHERE clinic_id=? AND status='In Progress' ORDER BY order_index LIMIT 1",
            [c['id']], one=True)
        if not current_construction:
            current_construction = query_db(
                "SELECT name FROM construction_tasks WHERE clinic_id=? AND status='Not Started' ORDER BY order_index LIMIT 1",
                [c['id']], one=True)
        construction_phase = current_construction['name'] if current_construction else 'Complete'

        total_expenses = query_db(
            "SELECT SUM(amount) as s FROM expenses WHERE clinic_id=?", [c['id']], one=True)['s'] or 0

        lease_net_cost_vip = None
        if lease:
            buildout = parse_dollar(lease['buildout_cost_estimate'])
            ti = parse_dollar(lease['ti_allowance'])
            if buildout is not None and ti is not None:
                lease_net_cost_vip = buildout - ti
            elif buildout is not None:
                lease_net_cost_vip = buildout

        bc = parse_dollar(c['buildout_cost']) if 'buildout_cost' in c.keys() else None
        ti_c = parse_dollar(c['ti_allowance']) if 'ti_allowance' in c.keys() else None
        net_cost_vip_clinic = None
        if bc is not None and ti_c is not None:
            net_cost_vip_clinic = bc - ti_c

        def fmt_money(n):
            if n is None:
                return ''
            return f'${n:,.0f}'

        all_data = {
            'clinic_name': c['name'],
            'state': c['state'] or '',
            'entity': c['entity'] or '',
            'site_status': c['site_status'] or '',
            'doctor_status': c['doctor_status'] or '',
            'targeting_opening_month': c['targeting_opening_month'] or '',
            'opening_date': c['opening_date'] or '',
            'setup_week': c['setup_week'] or '',
            'procedures_done_where': c['procedures_done_where'] or '',
            'paired_clinic': c['paired_clinic'] or '',
            'sg_project_manager': c['sg_project_manager'] or '',
            'sg_onsite_member': c['sg_onsite_member'] or '',
            'ops_onsite_member': c['ops_onsite_member'] or '',
            'lease_signed_date': c['lease_signed_date'] or '',
            'building_address': (lease['building_address'] if lease else '') or '',
            'suite_unit': (lease['suite_unit'] if lease else '') or '',
            'square_footage': (lease['square_footage'] if lease else '') or '',
            'landlord_name': (lease['landlord_name'] if lease else '') or '',
            'lease_start_date': (lease['lease_start_date'] if lease else '') or '',
            'lease_end_date': (lease['lease_end_date'] if lease else '') or '',
            'base_rent': (lease['base_rent'] if lease else '') or '',
            'cam_rent': (lease['cam_rent'] if lease else '') or '',
            'total_monthly_rent': (lease['total_monthly_rent'] if lease else '') or '',
            'annual_rent_increase_pct': (lease['annual_rent_increase_pct'] if lease else '') or '',
            'security_deposit': (lease['security_deposit'] if lease else '') or '',
            'ti_allowance': (lease['ti_allowance'] if lease else '') or '',
            'buildout_cost_estimate': (lease['buildout_cost_estimate'] if lease else '') or '',
            'lease_net_cost_vip': fmt_money(lease_net_cost_vip),
            'attorney_name': (lease['attorney_name'] if lease else '') or '',
            'overall_pct': f'{pct}%',
            'tasks_complete': str(done_tasks),
            'tasks_total': str(total_tasks),
            'tasks_overdue': str(overdue_tasks),
            'construction_phase': construction_phase,
            'supplies_pct': f'{supply_pct}%',
            'total_expenses': fmt_money(total_expenses),
            'buildout_cost': fmt_money(bc),
            'ti_allowance_clinic': fmt_money(ti_c),
            'net_cost_vip_clinic': fmt_money(net_cost_vip_clinic),
            'data_analysis_summary': c['data_analysis_summary'] or '',
            'clinic_notes': c['clinic_notes'] or '',
        }

        writer.writerow([all_data.get(f, '') for f in selected_fields])

    output = si.getvalue()
    return Response(output, mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=custom_report.csv'})


# ─── Due-date notifications check ────────────────────────────────────────────

@app.before_request
def check_due_notifications():
    if 'user_id' not in session:
        return
    if not request.endpoint or request.endpoint in ('static', 'notif_count'):
        return
    last_check = session.get('due_check_date')
    today = datetime.now().strftime('%Y-%m-%d')
    if last_check == today:
        return
    session['due_check_date'] = today
    uid = session['user_id']
    three_days = (datetime.now() + timedelta(days=3)).strftime('%Y-%m-%d')
    due_tasks = query_db(
        '''SELECT t.* FROM tasks t WHERE t.due_date <= ? AND t.due_date >= ?
           AND t.status NOT IN ('Complete') AND t.assignees LIKE ?''',
        [three_days, today, f'%{uid}%'])
    for t in due_tasks:
        existing = query_db(
            "SELECT id FROM notifications WHERE user_id=? AND message LIKE ? AND created_at >= date('now', '-1 day')",
            [uid, f'%{t["name"]}%'], one=True)
        if not existing:
            notify_user(uid, f'Task "{t["name"]}" is due soon ({t["due_date"]})', f'/task/{t["id"]}')

# ─── Main ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    with app.app_context():
        init_db()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
else:
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    with app.app_context():
        init_db()

@app.route('/clinic/<int:clinic_id>/set-lease-date', methods=['POST'])
@login_required
def set_lease_date(clinic_id):
    lease_date = request.form.get('lease_signed_date', '').strip()
    if lease_date:
        execute_db("UPDATE clinics SET lease_signed_date=? WHERE id=?", (lease_date, clinic_id))
        # Auto-set due dates for After Lease Signed tasks to lease_date + 14 days
        ld = datetime.strptime(lease_date, '%Y-%m-%d')
        due = (ld + timedelta(days=14)).strftime('%Y-%m-%d')
        execute_db(
            "UPDATE tasks SET due_date=? WHERE clinic_id=? AND time_phase='After Lease Signed' AND due_date IS NULL",
            (due, clinic_id)
        )
        flash('Lease signed date saved and task due dates updated!', 'success')
    return redirect(url_for('clinic_dashboard', clinic_id=clinic_id))
